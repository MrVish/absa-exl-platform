"""Build the program-tracking Excel for techno-functional + PM review.

6 sheets:
  1. Master Task List   (all ~85 tasks with full PM dimensions)
  2. Phase Summary      (rolled up by phase)
  3. Milestones & Gates (go/no-go checkpoints)
  4. Risk Register      (top risks + mitigations)
  5. Open Decisions     (the 'Need Comment' items + others)
  6. Dashboard          (counts by status / owner / phase, formula-driven)

Run from repo root: `uv run python scripts/build_program_excel.py`
Output: docs/absa-exl-program-plan.xlsx
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------- Styling primitives ----------

FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_BODY_BOLD = Font(name="Arial", size=10, bold=True)
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="1F3864")

FILL_HEADER = PatternFill("solid", start_color="1F3864")    # navy
FILL_BAND = PatternFill("solid", start_color="F2F2F2")      # zebra band

# Status color coding
FILL_DONE = PatternFill("solid", start_color="C8E6C9")      # green
FILL_PARTIAL = PatternFill("solid", start_color="FFF9C4")   # yellow
FILL_PLANNED = PatternFill("solid", start_color="FFE0B2")   # orange
FILL_BLOCKED = PatternFill("solid", start_color="FFCDD2")   # red
FILL_INPROG = PatternFill("solid", start_color="BBDEFB")    # blue

STATUS_FILLS = {
    "Done": FILL_DONE,
    "Partial": FILL_PARTIAL,
    "Planned": FILL_PLANNED,
    "Blocked": FILL_BLOCKED,
    "In Progress": FILL_INPROG,
    "Not Started": PatternFill("solid", start_color="EEEEEE"),
}

FILL_RISK_HIGH = PatternFill("solid", start_color="FFCDD2")
FILL_RISK_MED = PatternFill("solid", start_color="FFF9C4")
FILL_RISK_LOW = PatternFill("solid", start_color="C8E6C9")
RISK_FILLS = {"High": FILL_RISK_HIGH, "Medium": FILL_RISK_MED, "Low": FILL_RISK_LOW}

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------- Source plan: all tasks ----------
# Anchored to:
#   Month 1 = January 2026
#   Month 7 = July 2026 (steady state)

M1_START = date(2026, 1, 1)
M2_START = date(2026, 2, 1)
M3_START = date(2026, 3, 1)
M4_START = date(2026, 4, 1)
M5_START = date(2026, 5, 1)
M6_START = date(2026, 6, 1)
M7_START = date(2026, 7, 1)

# Phase windows (start, end)
PHASE_WINDOWS = {
    "P1": (M1_START, date(2026, 1, 31)),
    "P2": (date(2026, 1, 15), date(2026, 2, 28)),
    "P3": (date(2026, 1, 15), date(2026, 2, 28)),
    "P4": (M2_START, date(2026, 3, 31)),
    "P5": (M2_START, date(2026, 6, 30)),
    "P6": (M3_START, date(2026, 4, 30)),
    "P7": (M4_START, date(2026, 4, 30)),
    "P8a": (M5_START, date(2026, 6, 30)),
    "P8b": (M5_START, date(2026, 6, 30)),
    "P9a": (M7_START, date(2027, 1, 31)),
    "P9b": (M7_START, date(2027, 1, 31)),
    # Jenkins CI migration workstream — runs alongside Phase 4 (ADR-0011)
    "M1": (date(2026, 6, 8),  date(2026, 6, 14)),   # Foundation
    "M2": (date(2026, 6, 15), date(2026, 6, 28)),   # AWS-touching
    "M3": (date(2026, 6, 29), date(2026, 7, 5)),    # Cutover
}


# Tasks: (id, phase, category, task, owner, raci_r, raci_a, priority,
#         effort_days, deps, status, delivery, risk, notes)
TASKS: list[tuple[str, str, str, str, str, str, str, str, int, str, str, str, str, str]] = [
    # ---- Phase 1: Project Governance & Scope (Month 1) ----
    ("T001", "P1", "Project Governance & Scope", "Confirm project kickoff agenda and stakeholder list", "BOTH", "Program Mgr", "Steering Committee", "Critical", 2, "", "Not Started", "Planned", "Low", "Kickoff meeting"),
    ("T002", "P1", "Project Governance & Scope", "SPOCs from ABSA and EXL", "BOTH", "Program Mgr", "Exec Sponsor", "Critical", 1, "T001", "Not Started", "Planned", "Low", "Single Points of Contact each side"),
    ("T003", "P1", "Project Governance & Scope", "Finalize governance structure and working cadence", "BOTH", "Program Mgr", "Steering Committee", "Critical", 3, "T002", "Not Started", "Planned", "Low", "Weekly standup + monthly steering"),
    ("T004", "P1", "Project Governance & Scope", "Review scope, exclusions, assumptions, dependencies", "BOTH", "Program Mgr", "Steering Committee", "High", 3, "T003", "Not Started", "Planned", "Medium", "RACI + dependency log"),
    ("T005", "P1", "Project Governance & Scope", "Agree delivery plan for the initial 10 models", "BOTH", "Program Mgr", "ABSA Risk Owner", "Critical", 5, "T004", "Not Started", "Planned", "Medium", "Sequencing: Group 1 (2) -> Group 2 (8)"),
    ("T006", "P1", "Project Governance & Scope", "Confirm run frequency for each of the initial models", "BOTH", "Tech Lead", "ABSA Risk Owner", "High", 3, "T005", "Not Started", "Planned", "Low", "Daily / weekly / monthly cadence per model"),
    ("T007", "P1", "Project Governance & Scope", "Share approved model documents (initial set)", "ABSA", "ABSA Model Owner", "ABSA Model Owner", "Critical", 5, "T002", "Not Started", "Blocked", "High", "Model dev specs + validation reports"),
    ("T008", "P1", "Project Governance & Scope", "Share model code, supporting artefacts, benchmark outputs", "ABSA", "ABSA Model Owner", "ABSA Model Owner", "Critical", 5, "T007", "Not Started", "Blocked", "High", "Reference golden data for reconciliation"),
    ("T009", "P1", "Project Governance & Scope", "Agree PIR expectations and sign-off checkpoints", "BOTH", "Tech Lead", "ABSA Risk Owner", "Critical", 3, "T004", "Not Started", "Planned", "Medium", "Post-Implementation Review framework"),
    ("T010", "P1", "Project Governance & Scope", "Create project logger", "EXL", "Program Mgr", "Program Mgr", "Medium", 2, "T003", "Not Started", "Planned", "Low", "Jira / Confluence / equivalent"),
    # ---- Phase 2: Architecture, Access & Connectivity (M1-2) ----
    ("T011", "P2", "Architecture, Access & Connectivity", "Review target architecture + secure transfer approach", "BOTH", "Tech Lead", "Architecture Board", "Critical", 5, "T001", "In Progress", "Partial", "Medium", "ADRs 0001-0010 capture decisions"),
    ("T012", "P2", "Architecture, Access & Connectivity", "Approve final architecture for implementation", "ABSA", "ABSA Architect", "ABSA Architect", "Critical", 3, "T011", "Not Started", "Blocked", "High", "Awaits ABSA sign-off"),
    ("T013", "P2", "Architecture, Access & Connectivity", "Complete ABSA internal cloud approvals", "ABSA", "ABSA Architect", "ABSA Architect", "Critical", 10, "T012", "Not Started", "Blocked", "High", "ABSA-side governance"),
    ("T014", "P2", "Architecture, Access & Connectivity", "Provision EXL AWS landing zone accounts + baseline access", "EXL", "EXL Cloud Eng", "EXL Cloud Eng", "Critical", 5, "T013", "Not Started", "Partial", "Medium", "terraform/account-bootstrap/ ready"),
    ("T015", "P2", "Architecture, Access & Connectivity", "Set up IAM trust for cross-account connectivity", "BOTH", "EXL Cloud Eng", "ABSA Architect", "Critical", 3, "T014", "Not Started", "Partial", "Medium", "signing-foundation module wires policy"),
    ("T016", "P2", "Architecture, Access & Connectivity", "Complete IP whitelisting for approved endpoints", "ABSA", "ABSA Network", "ABSA Network", "High", 3, "T013", "Not Started", "Blocked", "Medium", "ABSA-side firewall rules"),
    ("T017", "P2", "Architecture, Access & Connectivity", "Share secure tunnel and connectivity details", "ABSA", "ABSA Network", "ABSA Network", "High", 2, "T016", "Not Started", "Blocked", "Medium", "VPC peering / TGW / PrivateLink decision"),
    ("T018", "P2", "Architecture, Access & Connectivity", "Configure encrypted S3 replication path", "BOTH", "EXL Cloud Eng", "ABSA Architect", "High", 5, "T015,T017", "Not Started", "Partial", "Medium", "s3-replication-{source,destination} modules ready"),
    ("T019", "P2", "Architecture, Access & Connectivity", "Validate end-to-end encrypted transfer", "BOTH", "EXL Cloud Eng", "ABSA Architect", "High", 3, "T018", "Not Started", "Planned", "Medium", "Real-account test"),
    ("T020", "P2", "Architecture, Access & Connectivity", "Finalize domain mapping for access endpoints", "BOTH", "EXL Cloud Eng", "ABSA Network", "Medium", 3, "T015", "Not Started", "Planned", "Low", "DNS / Route 53 plan"),
    ("T021", "P2", "Architecture, Access & Connectivity", "Configure SSO federation", "BOTH", "EXL Cloud Eng", "ABSA Identity", "High", 5, "T015", "Not Started", "Planned", "Medium", "Identity provider integration"),
    ("T022", "P2", "Architecture, Access & Connectivity", "Joint SSO testing + sign-off", "BOTH", "EXL Cloud Eng", "ABSA Identity", "High", 3, "T021", "Not Started", "Planned", "Medium", "After Phase 2 dependencies clear"),
    # ---- Phase 3: Data & Code Readiness (M1-2) ----
    ("T023", "P3", "Data & Code Readiness", "Confirm source datasets + file structure per model", "ABSA", "ABSA Data Owner", "ABSA Data Owner", "Critical", 5, "T005", "Not Started", "Blocked", "High", "Per-model data dictionary"),
    ("T024", "P3", "Data & Code Readiness", "Freeze input variables + requirements (initial set)", "ABSA", "ABSA Risk Owner", "ABSA Risk Owner", "Critical", 5, "T023", "Not Started", "Blocked", "High", "PIR mapping authority"),
    ("T025", "P3", "Data & Code Readiness", "Provide access to original dev environment", "ABSA", "ABSA Model Owner", "ABSA Model Owner", "High", 5, "T024", "Not Started", "Blocked", "Medium", "For code optimization (Phase 6)"),
    ("T026", "P3", "Data & Code Readiness", "Prepare scoring-ready data at agreed cadence", "EXL (ABSA Side)", "EXL Data Eng", "ABSA Data Owner", "High", 8, "T024", "Not Started", "Blocked", "Medium", "EXL operates inside ABSA's data plane"),
    ("T027", "P3", "Data & Code Readiness", "Prepare signed code package with documentation", "ABSA", "ABSA Model Owner", "ABSA Model Owner", "Critical", 8, "T008", "Not Started", "Blocked", "High", "Package contract spec is in repo (ADR-0010)"),
    ("T028", "P3", "Data & Code Readiness", "Transfer scoring-ready data, code, benchmarks to EXL", "ABSA", "ABSA Data Owner", "ABSA Data Owner", "Critical", 3, "T019,T026,T027", "Not Started", "Blocked", "Medium", "Cross-account S3 replication path"),
    ("T029", "P3", "Data & Code Readiness", "Validate receipt, completeness, integrity [Vishnu's Repo]", "EXL", "EXL Eng", "Tech Lead", "Critical", 2, "T028", "Done", "Done", "Low", "code-intake validate runs 5 checkers"),
    ("T030", "P3", "Data & Code Readiness", "Run data-quality checks: schema, volume, drift [Vishnu's Repo]", "EXL", "EXL Eng", "Tech Lead", "High", 3, "T029", "In Progress", "Partial", "Medium", "Schema + PIR coverage done; volume + drift TBD"),
    ("T031", "P3", "Data & Code Readiness", "Review and resolve data/code intake issues", "BOTH", "Tech Lead", "ABSA Data Owner", "High", 5, "T029,T030", "Not Started", "Planned", "Medium", "Triage process per model"),
    # ---- Phase 4: Platform Foundation Build (M2-3) ----
    ("T032", "P4", "Platform Foundation Build", "Build AWS landing zone + core networking", "EXL", "EXL Cloud Eng", "Tech Lead", "Critical", 10, "T014", "Not Started", "Partial", "Medium", "Modules ready; apply on real accounts"),
    ("T033", "P4", "Platform Foundation Build", "Configure IAM roles, KMS keys, secrets", "EXL", "EXL Cloud Eng", "Tech Lead", "Critical", 8, "T032", "Not Started", "Partial", "Medium", "signing-foundation module ready"),
    ("T034", "P4", "Platform Foundation Build", "Set up ML compute platform for model runs", "EXL", "EXL ML Eng", "Tech Lead", "Critical", 15, "T033", "Not Started", "Planned", "High", "Choice: SFN+Lambda vs SageMaker (decision needed)"),
    ("T035", "P4", "Platform Foundation Build", "Create model registry + version control standards", "EXL", "EXL Eng", "Tech Lead", "Critical", 5, "T033", "Done", "Done", "Low", "registry-api FastAPI + DynamoDB"),
    ("T036", "P4", "Platform Foundation Build", "Set up feature metadata + lineage capture", "EXL", "EXL Eng", "Tech Lead", "High", 8, "T035", "In Progress", "Partial", "Medium", "PIR captures input lineage; output TBD"),
    ("T037", "P4", "Platform Foundation Build", "Set up CI/CD pipelines for deployment + rollback", "EXL", "EXL DevOps", "Tech Lead", "High", 10, "T032", "In Progress", "Partial", "Medium", "GH Actions workflows in place"),
    ("T038", "P4", "Platform Foundation Build", "Create infrastructure-as-code templates", "EXL", "EXL Cloud Eng", "Tech Lead", "High", 5, "T032", "Done", "Done", "Low", "All terraform/ modules + per-env stacks"),
    ("T039", "P4", "Platform Foundation Build", "Set up observability, logging, alerts [Vishnu's Repo]", "EXL", "EXL DevOps", "Tech Lead", "High", 8, "T037", "In Progress", "Partial", "Medium", "Audit emission done; dashboards TBD"),
    ("T040", "P4", "Platform Foundation Build", "Set up audit trail + evidence capture", "EXL", "EXL Eng", "Compliance Lead", "Critical", 5, "T035", "Done", "Done", "Low", "registry-api emits append-only audit"),
    ("T041", "P4", "Platform Foundation Build", "Validate platform readiness vs security controls [Vishnu's Repo]", "BOTH", "Compliance Lead", "ABSA Risk Owner", "Critical", 5, "T040", "Done", "Done", "Low", "docs/compliance/control-matrix.md"),
    # ---- Phase 5: Controls, Security & Change Management (M2-6, cross-cutting) ----
    ("T042", "P5", "Controls, Security & Change Mgmt", "Define change-request workflow + approval matrix", "BOTH", "Compliance Lead", "ABSA Risk Owner", "High", 5, "T001", "In Progress", "Partial", "Medium", "registry-api approval state machine"),
    ("T043", "P5", "Controls, Security & Change Mgmt", "Set up review process for medium/high-risk changes [Vishnu's Repo]", "BOTH", "Compliance Lead", "ABSA Risk Owner", "High", 5, "T042", "In Progress", "Partial", "Medium", "approve/retire routes + audit log"),
    ("T044", "P5", "Controls, Security & Change Mgmt", "Define peer review + promotion gates for production", "BOTH", "Tech Lead", "Compliance Lead", "High", 3, "T042", "Done", "Done", "Low", "GH branch protection + drift gates + localstack-demo"),
    ("T045", "P5", "Controls, Security & Change Mgmt", "Document rollback requirements for planned changes", "EXL", "Tech Lead", "Compliance Lead", "Medium", 3, "T044", "In Progress", "Partial", "Medium", "Runbook structure exists; per-model TBD"),
    ("T046", "P5", "Controls, Security & Change Mgmt", "Set up release logging in agreed tracking tool", "EXL", "EXL DevOps", "Compliance Lead", "Medium", 3, "T040", "Done", "Done", "Low", "Audit log via registry-api"),
    ("T047", "P5", "Controls, Security & Change Mgmt", "Confirm data retention and archival rules", "BOTH", "Compliance Lead", "ABSA Risk Owner", "High", 3, "T042", "Not Started", "Planned", "Medium", "S3 lifecycle policies"),
    ("T048", "P5", "Controls, Security & Change Mgmt", "Confirm secure access controls for dashboard endpoints", "BOTH", "EXL Cloud Eng", "Compliance Lead", "High", 3, "T042", "Not Started", "Planned", "Medium", "After dashboard build (Phase 8)"),
    # ---- Phase 6: Code Optimization & Pipeline Setup (M3-4) ----
    ("T049", "P6", "Code Optimization & Pipeline Setup", "Review developer code for production readiness", "EXL", "EXL Eng", "Tech Lead", "Critical", 8, "T029,T034", "In Progress", "Partial", "Medium", "code-intake validate --strict"),
    ("T050", "P6", "Code Optimization & Pipeline Setup", "Standardize and optimize model scoring code", "EXL", "EXL ML Eng", "Tech Lead", "Critical", 15, "T025,T049", "Not Started", "Planned", "High", "Per-package work; uses dev env access"),
    ("T051", "P6", "Code Optimization & Pipeline Setup", "Validate scoring logic against dev benchmarks", "BOTH", "EXL ML Eng", "ABSA Model Owner", "Critical", 8, "T050", "Not Started", "Blocked", "High", "Reconciliation framework"),
    ("T052", "P6", "Code Optimization & Pipeline Setup", "Package scoring code for controlled deployment", "EXL", "EXL Eng", "Tech Lead", "Critical", 5, "T051", "Done", "Done", "Low", "Package contract spec (ADR-0010)"),
    ("T053", "P6", "Code Optimization & Pipeline Setup", "Create reusable scoring pipeline templates", "EXL", "EXL Eng", "Tech Lead", "Critical", 8, "T034,T052", "Done", "Done", "Low", "pipeline-factory ASL templates"),
    ("T054", "P6", "Code Optimization & Pipeline Setup", "Register models + pipeline versions in registry", "EXL", "EXL Eng", "Tech Lead", "High", 3, "T035,T053", "Done", "Done", "Low", "register-pipeline CLI (SigV4)"),
    ("T055", "P6", "Code Optimization & Pipeline Setup", "Set up model run schedules for approved cadence", "BOTH", "EXL DevOps", "ABSA Risk Owner", "High", 3, "T006,T054", "Not Started", "Planned", "Medium", "EventBridge schedules"),
    ("T056", "P6", "Code Optimization & Pipeline Setup", "Complete code intake validation + metadata capture [Vishnu's Repo]", "EXL", "EXL Eng", "Tech Lead", "Critical", 3, "T049", "Done", "Done", "Low", "code-intake generate-manifest"),
    # ---- Phase 7: Group 1 (Month 4) ----
    ("T057", "P7", "Group 1: First 2 Models", "Prepare Group 1 onboarding plan", "EXL", "Program Mgr", "Tech Lead", "Critical", 5, "T056", "Not Started", "Planned", "Medium", "Per-model checklist"),
    ("T058", "P7", "Group 1: First 2 Models", "Run initial scoring for Group 1 models", "EXL", "EXL ML Eng", "Tech Lead", "Critical", 5, "T057", "Not Started", "Blocked", "High", "Needs real ABSA accounts + data"),
    ("T059", "P7", "Group 1: First 2 Models", "Reconcile Group 1 outputs with benchmarks", "BOTH", "EXL ML Eng", "ABSA Model Owner", "Critical", 5, "T058", "Not Started", "Blocked", "High", "Byte/value comparison vs ABSA golden"),
    ("T060", "P7", "Group 1: First 2 Models", "Resolve defects identified during initial output testing", "BOTH", "EXL ML Eng", "ABSA Model Owner", "Critical", 8, "T059", "Not Started", "Planned", "High", "Defect log + retest cadence"),
    ("T061", "P7", "Group 1: First 2 Models", "Complete PIR for Group 1 models", "BOTH", "ABSA Risk Owner", "ABSA Risk Owner", "Critical", 5, "T060", "Not Started", "Planned", "High", "Post-Implementation Review"),
    ("T062", "P7", "Group 1: First 2 Models", "MILESTONE: Obtain sign-off for Group 1", "ABSA", "ABSA Risk Owner", "ABSA Risk Owner", "Critical", 2, "T061", "Not Started", "Blocked", "High", "*** GATE TO GROUP 2 ***"),
    # ---- Phase 8a: Group 2 (M5-6) ----
    ("T063", "P8a", "Group 2: Remaining 8 Models", "Prepare onboarding plan for Group 2", "EXL", "Program Mgr", "Tech Lead", "Critical", 5, "T062", "Not Started", "Planned", "Medium", "Reuses Group 1 lessons + templates"),
    ("T064", "P8a", "Group 2: Remaining 8 Models", "Run onboarding and scoring for Group 2", "EXL", "EXL ML Eng", "Tech Lead", "Critical", 20, "T063", "Not Started", "Planned", "Medium", "Each uses pipeline template"),
    ("T065", "P8a", "Group 2: Remaining 8 Models", "Reconcile Group 2 outputs with benchmarks", "BOTH", "EXL ML Eng", "ABSA Model Owner", "Critical", 10, "T064", "Not Started", "Planned", "Medium", "Same framework as Group 1"),
    ("T066", "P8a", "Group 2: Remaining 8 Models", "Resolve defects during Group 2 testing", "BOTH", "EXL ML Eng", "ABSA Model Owner", "Critical", 10, "T065", "Not Started", "Planned", "Medium", ""),
    ("T067", "P8a", "Group 2: Remaining 8 Models", "Complete PIR for Group 2 models", "BOTH", "ABSA Risk Owner", "ABSA Risk Owner", "Critical", 8, "T066", "Not Started", "Planned", "Medium", ""),
    ("T068", "P8a", "Group 2: Remaining 8 Models", "MILESTONE: Obtain sign-off for Group 2", "ABSA", "ABSA Risk Owner", "ABSA Risk Owner", "Critical", 2, "T067", "Not Started", "Planned", "Medium", "Closes initial-onboarding scope"),
    # ---- Phase 8b: Dashboards & Delivery (M5-6, parallel with Group 2) ----
    ("T069", "P8b", "Dashboards & Delivery", "Agree dashboard views, recipients, access levels", "BOTH", "Program Mgr", "ABSA Risk Owner", "High", 3, "T048", "Not Started", "Planned", "Low", "Who sees what, when"),
    ("T070", "P8b", "Dashboards & Delivery", "Build monitoring dashboard for runs + exceptions", "EXL", "EXL DevOps", "Tech Lead", "High", 10, "T069", "Not Started", "Planned", "Medium", "CloudWatch + Grafana or equivalent"),
    ("T071", "P8b", "Dashboards & Delivery", "Configure notifications + alert routing for failures", "BOTH", "EXL DevOps", "Tech Lead", "High", 3, "T070", "Not Started", "Planned", "Medium", "SNS / PagerDuty / Slack"),
    ("T072", "P8b", "Dashboards & Delivery", "Validate dashboard for inputs, outputs, failures", "BOTH", "EXL DevOps", "ABSA Risk Owner", "High", 3, "T070", "Not Started", "Planned", "Low", ""),
    ("T073", "P8b", "Dashboards & Delivery", "Set up secure output delivery process to ABSA", "EXL", "EXL Cloud Eng", "Compliance Lead", "Critical", 5, "T018", "In Progress", "Partial", "Medium", "S3 replication path"),
    ("T074", "P8b", "Dashboards & Delivery", "Confirm operational runbooks + support contacts", "BOTH", "Tech Lead", "Program Mgr", "High", 3, "T070", "In Progress", "Partial", "Medium", "docs/runbooks/ started"),
    ("T075", "P8b", "Dashboards & Delivery", "Publish runbooks + operating procedures + handover", "EXL", "Tech Lead", "Program Mgr", "Critical", 5, "T074", "In Progress", "Partial", "Medium", "Demo + key-rotation runbooks done"),
    # ---- Phase 9a: Steady State Support (M7+) ----
    ("T076", "P9a", "Steady State Support", "Start scheduled production scoring", "EXL", "EXL DevOps", "Tech Lead", "Critical", 5, "T068,T075", "Not Started", "Planned", "Medium", ""),
    ("T077", "P9a", "Steady State Support", "Monitor daily, weekly, monthly scoring runs", "EXL", "EXL DevOps", "Tech Lead", "High", 0, "T076", "Not Started", "Planned", "Low", "Ongoing"),
    ("T078", "P9a", "Steady State Support", "Track run failures + coordinate recovery actions", "BOTH", "EXL DevOps", "ABSA Risk Owner", "Critical", 0, "T076", "Not Started", "Planned", "Medium", "Incident response process"),
    ("T079", "P9a", "Steady State Support", "Maintain audit trail for every scoring run and delivery", "EXL", "EXL Eng", "Compliance Lead", "Critical", 0, "T076", "Done", "Done", "Low", "registry-api audit + signed manifests"),
    ("T080", "P9a", "Steady State Support", "Review data drift, anomalies, scoring exceptions", "BOTH", "EXL ML Eng", "ABSA Risk Owner", "High", 0, "T076", "Not Started", "Planned", "Medium", "*** Need Comment: drift framework ***"),
    ("T081", "P9a", "Steady State Support", "Manage approved changes", "BOTH", "Compliance Lead", "ABSA Risk Owner", "High", 0, "T076", "Not Started", "Planned", "Medium", "*** Need Comment: change-mgmt cadence ***"),
    ("T082", "P9a", "Steady State Support", "Refresh access, whitelists, secrets when required", "BOTH", "EXL Cloud Eng", "Compliance Lead", "High", 0, "T076", "Not Started", "Planned", "Medium", "*** Need Comment: rotation cadence ***"),
    ("T083", "P9a", "Steady State Support", "Archive model data + outputs as per retention rules", "EXL", "EXL Cloud Eng", "Compliance Lead", "Medium", 0, "T076", "Not Started", "Planned", "Low", "S3 lifecycle policies"),
    ("T084", "P9a", "Steady State Support", "Conduct periodic service review with ABSA", "BOTH", "Program Mgr", "ABSA Risk Owner", "High", 0, "T076", "Not Started", "Planned", "Low", "Quarterly cadence"),
    ("T085", "P9a", "Steady State Support", "Handle operational incidents + secondary support", "EXL", "EXL DevOps", "Program Mgr", "Critical", 0, "T076", "Not Started", "Planned", "Medium", ""),
    # ---- Phase 9b: New Model Onboarding (M7+) ----
    ("T086", "P9b", "New Model Onboarding", "Review each new onboarding request + classify model type", "BOTH", "Tech Lead", "Program Mgr", "High", 2, "T068", "Not Started", "Planned", "Low", "Standard-batch / Scalable-batch / Realtime"),
    ("T087", "P9b", "New Model Onboarding", "Collect approved artefacts, inputs, benchmark outputs", "ABSA", "ABSA Model Owner", "ABSA Model Owner", "High", 5, "T086", "Not Started", "Planned", "Medium", "Same intake contract"),
    ("T088", "P9b", "New Model Onboarding", "Reuse existing template for subsequent model of same type", "EXL", "EXL Eng", "Tech Lead", "Medium", 3, "T087", "Done", "Done", "Low", "pipeline-factory regenerates"),
    ("T089", "P9b", "New Model Onboarding", "Build type-specific controls for new model category", "EXL", "EXL Eng", "Tech Lead", "Medium", 10, "T087", "Not Started", "Planned", "Medium", "If new tier introduced"),
    ("T090", "P9b", "New Model Onboarding", "Run validation, scoring, PIR for each newly onboarded model", "BOTH", "EXL ML Eng", "ABSA Model Owner", "High", 10, "T088", "Not Started", "Planned", "Medium", "Same Group 1/2 framework"),
    ("T091", "P9b", "New Model Onboarding", "Add approved new models to registry + run schedule", "EXL", "EXL Eng", "Tech Lead", "Medium", 2, "T090", "Done", "Done", "Low", "register-pipeline + EventBridge"),
    # ---- Jenkins CI Migration (ADR-0011), runs alongside Phase 4 ----
    ("T101", "M1", "Jenkins Migration: Foundation",  "ADR-0011 drafted + committed",                              "EXL", "Tech Lead",      "Tech Lead",       "Critical", 1, "",       "Done",        "Done",    "Low",    "Status: Proposed"),
    ("T102", "M1", "Jenkins Migration: Foundation",  "Shared library 'absa-ci' skeleton + 5 example Jenkinsfiles", "EXL", "EXL DevOps",   "Tech Lead",       "Critical", 2, "T101",   "Done",        "Done",    "Low",    "ci/jenkins/ committed"),
    ("T103", "M1", "Jenkins Migration: Foundation",  "ABSA confirms Jenkins identity model (IRSA on EKS preferred)", "ABSA", "ABSA Cloud Platform", "ABSA Architect", "Critical", 3, "T101", "Not Started", "Blocked", "High",   "*** GATING — unblocks every M2 task ***"),
    ("T104", "M1", "Jenkins Migration: Foundation",  "Stand up Jenkins sandbox + register absa-ci shared library","EXL", "EXL DevOps",   "Tech Lead",       "Critical", 5, "T103",   "Not Started", "Blocked", "Medium", "Sandbox proves library wiring before any real workflow flips"),
    ("T105", "M1", "Jenkins Migration: Foundation",  "Run python-validate.Jenkinsfile end-to-end in sandbox",     "EXL", "EXL DevOps",   "Tech Lead",       "High",     2, "T104",   "Not Started", "Blocked", "Medium", "Proving ground — confirms commit-status reporting works"),
    ("T106", "M1", "Jenkins Migration: Foundation",  "Stamp ADR-0011 Status: Accepted after sandbox proves out",  "EXL", "Tech Lead",     "Tech Lead",       "High",     1, "T105",   "Not Started", "Blocked", "Low",    ""),
    ("T107", "M2", "Jenkins Migration: AWS-touching","signing-foundation Terraform: identity_provider variable",  "EXL", "EXL Cloud Eng", "Tech Lead",       "Critical", 3, "T103",   "Not Started", "Blocked", "Medium", "Trust both providers during cutover"),
    ("T108", "M2", "Jenkins Migration: AWS-touching","Apply signing-foundation update to exl-prod stack",         "EXL", "EXL Cloud Eng", "Tech Lead",       "Critical", 1, "T107",   "Not Started", "Blocked", "Medium", ""),
    ("T109", "M2", "Jenkins Migration: AWS-touching","Port publish-signing-key.yml -> Jenkinsfile",               "EXL", "EXL DevOps",     "Tech Lead",       "High",     2, "T108",   "Not Started", "Blocked", "Low",    "Held in M1 — identity model decided here"),
    ("T110", "M2", "Jenkins Migration: AWS-touching","Validate pipeline-factory.Jenkinsfile against real AWS",    "EXL", "EXL DevOps",     "Tech Lead",       "Critical", 3, "T108",   "Not Started", "Blocked", "High",   "Signature byte-equivalence GHA vs Jenkins"),
    ("T111", "M2", "Jenkins Migration: AWS-touching","Parallel-run GHA + Jenkins for 1 week",                     "EXL", "EXL DevOps",     "Tech Lead",       "High",     5, "T110",   "Not Started", "Blocked", "Medium", "Compare digests, signatures, manifests byte-by-byte"),
    ("T112", "M3", "Jenkins Migration: Cutover",     "Flip GitHub branch protection to Jenkins commit statuses",  "EXL", "EXL DevOps",     "Tech Lead",       "Critical", 1, "T111",   "Not Started", "Blocked", "High",   "One-way step; rollback = re-add GHA contexts"),
    ("T113", "M3", "Jenkins Migration: Cutover",     "Disable GHA workflows (.github/workflows/*.disabled.yml)",  "EXL", "EXL DevOps",     "Tech Lead",       "High",     1, "T112",   "Not Started", "Blocked", "Low",    "Keep one release for rollback safety"),
    ("T114", "M3", "Jenkins Migration: Cutover",     "Sweep ADR-0003, ADR-0009, runbooks, technical-overview",    "EXL", "Tech Lead",      "Tech Lead",       "Medium",   2, "T112",   "Not Started", "Blocked", "Low",    "ADRs cross-referenced in M1; full sweep here"),
    ("T115", "M3", "Jenkins Migration: Cutover",     "Promote compliance matrix Phase 3 rows from Proposed -> Live", "EXL", "Compliance Lead", "Compliance Lead", "High",  1, "T112",   "Not Started", "Blocked", "Low",    "Auditor evidence URLs point to Jenkins builds"),
]


def _phase_dates(phase_id: str, idx: int, total: int) -> tuple[date, date]:
    """Spread a task across its phase window proportionally by index/total."""
    start_window, end_window = PHASE_WINDOWS[phase_id]
    span = (end_window - start_window).days
    if total <= 1:
        return start_window, end_window
    chunk = max(span // total, 3)
    s = start_window + timedelta(days=int(idx * (span / total)))
    e = s + timedelta(days=chunk)
    if e > end_window:
        e = end_window
    return s, e


# ---------- Sheet builders ----------


def _format_header_row(ws, row_idx: int, columns: list[str]) -> None:
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def _set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_master_task_list(wb: Workbook) -> None:
    ws = wb.create_sheet("Master Task List")
    ws.sheet_view.showGridLines = False

    # Title row
    ws["A1"] = "ABSA x EXL Model Hosting — Master Task List"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:O1")
    ws.row_dimensions[1].height = 24

    headers = [
        "Task ID", "Phase", "Category", "Task",
        "Owner", "RACI: R", "RACI: A",
        "Priority", "Effort (d)",
        "Start", "End", "Duration (d)",
        "Dependencies", "Status", "Delivery", "Risk", "Notes",
    ]
    _format_header_row(ws, 3, headers)

    # Track phase indices for date spreading
    phase_counts: dict[str, int] = {}
    for t in TASKS:
        phase_counts.setdefault(t[1], 0)
        phase_counts[t[1]] += 1

    phase_indices: dict[str, int] = {}
    row = 4
    for task in TASKS:
        (task_id, phase, category, name, owner, raci_r, raci_a, priority,
         effort_days, deps, status, delivery, risk, notes) = task

        phase_indices.setdefault(phase, 0)
        idx = phase_indices[phase]
        phase_indices[phase] += 1
        total = phase_counts[phase]
        start_dt, end_dt = _phase_dates(phase, idx, total)

        values = [
            task_id, phase, category, name,
            owner, owner, raci_a,
            priority, effort_days,
            start_dt, end_dt,
            f"=K{row}-J{row}+1",
            deps, status, delivery, risk, notes,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_THIN
            if isinstance(val, date):
                cell.number_format = "yyyy-mm-dd"

        # Status fill on the Delivery column
        delivery_cell = ws.cell(row=row, column=15)
        if delivery in STATUS_FILLS:
            delivery_cell.fill = STATUS_FILLS[delivery]

        # Status fill on the Status column
        status_cell = ws.cell(row=row, column=14)
        if status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[status]

        # Risk fill
        risk_cell = ws.cell(row=row, column=16)
        if risk in RISK_FILLS:
            risk_cell.fill = RISK_FILLS[risk]

        # Priority styling
        priority_cell = ws.cell(row=row, column=8)
        if priority == "Critical":
            priority_cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
        elif priority == "High":
            priority_cell.font = Font(name="Arial", size=10, bold=True, color="ED7D31")

        # Milestone row emphasis
        if "MILESTONE" in name:
            for col_idx in range(1, len(values) + 1):
                ws.cell(row=row, column=col_idx).font = FONT_BODY_BOLD

        row += 1

    # Column widths
    _set_col_widths(ws, {
        "A": 10, "B": 7, "C": 32, "D": 60,
        "E": 16, "F": 14, "G": 20,
        "H": 11, "I": 10,
        "J": 12, "K": 12, "L": 12,
        "M": 16, "N": 14, "O": 12, "P": 9, "Q": 50,
    })
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "E4"
    ws.auto_filter.ref = f"A3:Q{row - 1}"


def build_phase_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Phase Summary"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 24

    headers = [
        "Phase", "Months", "Title",
        "Task count (formula)", "Effort total (formula)",
        "Critical milestones", "Status snapshot",
    ]
    _format_header_row(ws, 3, headers)

    phases = [
        ("P1", "M1",      "Project Governance & Scope",        "SPOCs identified · scope sign-off"),
        ("P2", "M1-2",    "Architecture, Access & Connectivity", "ABSA architecture sign-off · IAM trust established"),
        ("P3", "M1-2",    "Data & Code Readiness",             "First model code + data + benchmarks in intake pipeline"),
        ("P4", "M2-3",    "Platform Foundation Build",         "Real AWS accounts deployed via Terraform"),
        ("P5", "M2-6",    "Controls, Security & Change Mgmt",  "Cross-cutting; runs alongside Phases 4-8"),
        ("P6", "M3-4",    "Code Optimization & Pipeline Setup","Pipeline template registered for first model"),
        ("P7", "M4",      "Group 1: First 2 Models",           "*** ABSA SIGN-OFF — BIGGEST MILESTONE ***"),
        ("P8a","M5-6",    "Group 2: Remaining 8 Models",       "All 10 models in production scoring"),
        ("P8b","M5-6",    "Dashboards & Delivery Readiness",   "Dashboards live; runbooks published"),
        ("P9a","M7+",     "Steady State Support",              "Quarterly reviews; ongoing ops"),
        ("P9b","M7+",     "New Model Onboarding",              "Template-driven new model ramp"),
        ("M1", "Jun W2",  "Jenkins Migration: Foundation",     "ADR-0011 accepted · sandbox proves library"),
        ("M2", "Jun W3-4","Jenkins Migration: AWS-touching",   "signing-foundation supports both providers · parallel run"),
        ("M3", "Jul W1",  "Jenkins Migration: Cutover",        "Branch protection flipped · GHA retired"),
    ]
    row = 4
    for phase_id, months, title, milestone in phases:
        # Formulas counting tasks in this phase
        task_count_formula = f"=COUNTIF('Master Task List'!B:B,\"{phase_id}\")"
        effort_formula = f"=SUMIF('Master Task List'!B:B,\"{phase_id}\",'Master Task List'!I:I)"

        status_formula = (
            f'=COUNTIFS(\'Master Task List\'!B:B,"{phase_id}",\'Master Task List\'!O:O,"Done")&" Done · "&'
            f'COUNTIFS(\'Master Task List\'!B:B,"{phase_id}",\'Master Task List\'!O:O,"Partial")&" Partial · "&'
            f'COUNTIFS(\'Master Task List\'!B:B,"{phase_id}",\'Master Task List\'!O:O,"Planned")&" Planned · "&'
            f'COUNTIFS(\'Master Task List\'!B:B,"{phase_id}",\'Master Task List\'!O:O,"Blocked")&" Blocked"'
        )

        values = [
            phase_id, months, title,
            task_count_formula, effort_formula,
            milestone, status_formula,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY_BOLD if phase_id == "P7" else FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_THIN
        if phase_id == "P7":
            for col_idx in range(1, len(values) + 1):
                ws.cell(row=row, column=col_idx).fill = FILL_BLOCKED
        row += 1

    _set_col_widths(ws, {"A": 8, "B": 8, "C": 36, "D": 18, "E": 18, "F": 50, "G": 50})
    ws.row_dimensions[3].height = 30


def build_milestones(wb: Workbook) -> None:
    ws = wb.create_sheet("Milestones & Gates")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Milestones & Gates"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 24

    headers = ["Milestone ID", "Description", "Target Date", "Owner", "Status", "Gate Criteria"]
    _format_header_row(ws, 3, headers)

    milestones = [
        ("M01", "Project kickoff complete + SPOCs identified",        date(2026, 1, 15), "Program Mgr",       "Not Started", "Signed kickoff minutes; SPOC names on file"),
        ("M02", "Scope + 10-model delivery plan signed off",          date(2026, 1, 31), "Steering Committee","Not Started", "Scope doc approved; sequencing locked"),
        ("M03", "Architecture sign-off from ABSA",                    date(2026, 2, 15), "ABSA Architect",    "Blocked",     "Architecture review board minutes + decision log"),
        ("M04", "EXL AWS landing zone live + IAM trust established",  date(2026, 2, 28), "EXL Cloud Eng",     "Partial",     "Real account IDs provisioned; cross-account roles verifiable"),
        ("M05", "First ABSA package transferred + validated",         date(2026, 2, 28), "Tech Lead",         "Blocked",     "code-intake validate exits 0; manifest signed"),
        ("M06", "Platform foundation built + security control sign-off", date(2026, 3, 31), "Tech Lead",      "Partial",     "Control matrix evidence + readiness review"),
        ("M07", "Code optimization complete + pipeline templates live", date(2026, 4, 15), "Tech Lead",       "Planned",     "Pipeline template registered for first model type"),
        ("M08", "*** GROUP 1 SIGN-OFF (2 MODELS) ***",                date(2026, 4, 30), "ABSA Risk Owner",   "Blocked",     "PIR complete; benchmark reconciliation passes"),
        ("M09", "*** GROUP 2 SIGN-OFF (REMAINING 8 MODELS) ***",      date(2026, 6, 30), "ABSA Risk Owner",   "Planned",     "All 10 models reconciled; PIR sign-off"),
        ("M10", "Dashboards + delivery handover published",           date(2026, 6, 30), "Program Mgr",       "Planned",     "Runbooks published; support contacts confirmed"),
        ("M11", "Steady-state go-live",                               date(2026, 7, 1),  "Program Mgr",       "Planned",     "First quarterly review scheduled; ops contacts staffed"),
        ("M12", "Jenkins identity model confirmed (IRSA vs instance)",date(2026, 6, 14), "ABSA Cloud Platform","Blocked",   "*** Gates entire Sprint M2 of Jenkins migration ***"),
        ("M13", "GHA retired; Jenkins is the sole CI gate",           date(2026, 7, 5),  "EXL DevOps",        "Planned",     "Branch protection flipped; .disabled.yml retained 1 release"),
    ]
    row = 4
    for m in milestones:
        for col_idx, val in enumerate(m, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY_BOLD if "GROUP" in m[1] else FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_THIN
            if isinstance(val, date):
                cell.number_format = "yyyy-mm-dd"
        status_cell = ws.cell(row=row, column=5)
        if m[4] in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[m[4]]
        if "GROUP" in m[1]:
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).fill = FILL_BLOCKED
        row += 1

    _set_col_widths(ws, {"A": 14, "B": 50, "C": 13, "D": 22, "E": 14, "F": 50})
    ws.row_dimensions[3].height = 30


def build_risk_register(wb: Workbook) -> None:
    ws = wb.create_sheet("Risk Register")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Risk Register"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 24

    headers = ["Risk ID", "Description", "Phase", "Probability (1-5)",
               "Impact (1-5)", "Score (P*I)", "Mitigation", "Owner"]
    _format_header_row(ws, 3, headers)

    risks = [
        ("R01", "ABSA account onboarding (real AWS account IDs + IAM principals) delayed", "P2/P4", 4, 5, "Pre-stage Terraform + use LocalStack demo to keep CI green; escalate to Steering by end of Month 1", "ABSA Architect"),
        ("R02", "ABSA-side internal cloud approvals slip past Month 2", "P2", 3, 5, "Start approval process before kickoff; weekly check-in with ABSA cloud team", "ABSA Architect"),
        ("R03", "Group 1 sign-off (Month 4) slips, cascading into Group 2 and steady state", "P7", 3, 5, "Use Group 1 onboarding plan as dress rehearsal; pre-stage Group 2 templates", "ABSA Risk Owner"),
        ("R04", "ML compute platform choice (SFN+Lambda vs SageMaker) blocks Phase 4 build", "P4", 3, 4, "Decision required by end of Month 2; architecture review board agenda item", "Tech Lead"),
        ("R05", "Benchmark reconciliation discrepancies require model code rewrite", "P6/P7", 3, 4, "Validate scoring logic incrementally during code optimization; record per-feature deltas", "EXL ML Eng"),
        ("R06", "Cross-account encrypted transfer path not validated end-to-end before Phase 4", "P2", 2, 4, "Phase-gate Phase 4 on Phase 2 transfer validation; LocalStack demo verifies the shape today", "EXL Cloud Eng"),
        ("R07", "SSO federation testing reveals incompatible identity provider config", "P2", 2, 3, "Joint test in Month 2; fall back to direct IAM if SSO blocks", "EXL Cloud Eng"),
        ("R08", "Data quality issues (drift, schema mismatch) in initial intake delay code optimization", "P3/P6", 3, 4, "Run code-intake validate on first arrival; reject + return to ABSA immediately", "EXL Eng"),
        ("R09", "PIR review reveals input variable misalignment, requires re-scope", "P7", 2, 5, "Lock input variables in Phase 1; PIR rehearsal during Group 1", "ABSA Risk Owner"),
        ("R10", "Production scoring SLA cannot be met on selected ML compute platform", "P9a", 2, 4, "Load-test in Phase 7 before Group 1 go-live", "EXL ML Eng"),
        ("R11", "KMS signing key compromise during operations", "P9a", 1, 5, "Rotation runbook (docs/runbooks/kms-key-rotation.md) in place; quarterly drill", "Compliance Lead"),
        ("R12", "Change management workflow not finalized before first production change", "P5", 2, 4, "Use registry-api approval state machine as baseline; document workflow by Month 3", "Compliance Lead"),
        ("R13", "ABSA does not run Jenkins on EKS — IRSA path closed, fall back to EC2 instance profile",       "M1", 3, 3, "ADR-0011 already documents the EC2 fallback path; awsLogin step handles both modes; identified as a known fallback rather than a blocker", "Tech Lead"),
        ("R14", "Jenkins shared library Groovy has runtime bugs that only surface against a real Jenkins instance", "M1", 3, 3, "Sprint M1 sandbox proves library before Sprint M2 starts; 5 example Jenkinsfiles share the same shared-library surface so bugs fail consistently", "EXL DevOps"),
        ("R15", "Signature byte-equivalence fails between GHA and Jenkins runs (canonical_json sensitivity to env)", "M2", 2, 5, "Parallel-run week explicitly compares manifest digests; canonical_json is deterministic by spec but Python version drift could shift behaviour", "Tech Lead"),
    ]
    row = 4
    for r in risks:
        (rid, desc, phase, prob, impact, mitigation, owner) = r
        values = [
            rid, desc, phase, prob, impact,
            f"=D{row}*E{row}",
            mitigation, owner,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_THIN

        # Score color
        score = prob * impact
        score_cell = ws.cell(row=row, column=6)
        if score >= 15:
            score_cell.fill = FILL_RISK_HIGH
            score_cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
        elif score >= 8:
            score_cell.fill = FILL_RISK_MED
            score_cell.font = Font(name="Arial", size=10, bold=True, color="ED7D31")
        else:
            score_cell.fill = FILL_RISK_LOW

        row += 1

    _set_col_widths(ws, {"A": 8, "B": 50, "C": 9, "D": 13, "E": 12, "F": 13, "G": 60, "H": 22})
    ws.row_dimensions[3].height = 30


def build_open_decisions(wb: Workbook) -> None:
    ws = wb.create_sheet("Open Decisions")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Open Decisions"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 24

    headers = ["Decision ID", "Description", "Phase Impact", "Owner", "Due Date", "Impact if not decided"]
    _format_header_row(ws, 3, headers)

    decisions = [
        ("D01", "Data drift detection framework: which metrics (PSI / KL-divergence / feature-distribution shift)? What window? Manual review vs automated alerting?", "P9a", "ABSA Risk Owner + EXL ML Eng", date(2026, 5, 31), "Cannot ship dashboards (Phase 8b) without agreed drift metrics. Steady state ops lacks anomaly framework."),
        ("D02", "Change management cadence: production-change approval — registry's approve/retire state machine or separate Jira workflow?", "P5/P9a", "Compliance Lead + ABSA Risk Owner", date(2026, 3, 31), "Group 1 production releases blocked without an agreed change workflow."),
        ("D03", "Access / secret rotation cadence (IAM principals, registrar tokens, dashboard access). KMS signing key has a runbook; the rest don't.", "P5/P9a", "Compliance Lead", date(2026, 4, 30), "Audit findings if rotation cadence not documented before steady state."),
        ("D04", "ML compute platform: Step Functions + Lambda OR SageMaker Pipelines?", "P4", "Tech Lead + Architecture Board", date(2026, 2, 28), "Blocks Phase 4 ML compute setup. Affects pipeline-factory ASL renderer scope."),
        ("D05", "Real-time scoring tier SLA: what's the response budget? Lambda+APIGW / ECS Fargate / SageMaker endpoint?", "P4", "Tech Lead + ABSA Architect", date(2026, 3, 31), "Realtime ASL template is currently a placeholder. SLA drives the implementation choice."),
        ("D06", "Data-movement choice: S3 cross-account replication or SFTP transfer for scoring inputs/outputs?", "P2", "ABSA Architect + EXL Cloud Eng", date(2026, 2, 15), "Blocks Phase 2 secure transfer build. Both directions need this lock."),
        ("D07", "Multi-region: is the platform single-region (eu-west-1) or multi-region from day 1?", "P4", "ABSA Architect", date(2026, 3, 15), "Affects KMS replica strategy, signing-foundation scope, runbook structure."),
        ("D08", "Dashboard hosting: Grafana / CloudWatch native / QuickSight / something ABSA-supplied?", "P8b", "Program Mgr + ABSA Risk Owner", date(2026, 4, 30), "Dashboard build (Phase 8b) blocked without tool decision."),
        ("D09", "Output delivery format to ABSA: signed S3 bucket / SFTP push / API pull?", "P8b", "ABSA Architect + EXL Cloud Eng", date(2026, 4, 30), "Cannot finalize delivery runbook."),
    ]
    row = 4
    for d in decisions:
        for col_idx, val in enumerate(d, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_THIN
            if isinstance(val, date):
                cell.number_format = "yyyy-mm-dd"
        row += 1

    _set_col_widths(ws, {"A": 10, "B": 60, "C": 10, "D": 30, "E": 13, "F": 55})
    ws.row_dimensions[3].height = 30


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Program Dashboard"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 26

    # ---- Section A: counts by status ----
    ws["A3"] = "Tasks by Delivery Status"
    ws["A3"].font = FONT_BODY_BOLD
    _format_header_row(ws, 4, ["Status", "Count"])
    statuses = ["Done", "Partial", "Planned", "Blocked", "In Progress", "Not Started"]
    row = 5
    for s in statuses:
        ws.cell(row=row, column=1, value=s).font = FONT_BODY
        ws.cell(row=row, column=2, value=f'=COUNTIF(\'Master Task List\'!O:O,"{s}")')
        ws.cell(row=row, column=1).fill = STATUS_FILLS.get(s, PatternFill())
        for c in range(1, 3):
            ws.cell(row=row, column=c).border = BORDER_THIN
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="Total").font = FONT_BODY_BOLD
    ws.cell(row=row, column=2, value=f"=SUM(B5:B{row - 1})").font = FONT_BODY_BOLD
    for c in range(1, 3):
        ws.cell(row=row, column=c).border = BORDER_THIN

    # ---- Section B: counts by owner ----
    ws["D3"] = "Tasks by Owner"
    ws["D3"].font = FONT_BODY_BOLD
    _format_header_row(ws, 4, ["Owner", "Count"])
    ws.cell(row=4, column=4).value = "Owner"
    ws.cell(row=4, column=5).value = "Count"
    owners = ["BOTH", "EXL", "ABSA", "EXL (ABSA Side)"]
    row = 5
    for o in owners:
        ws.cell(row=row, column=4, value=o).font = FONT_BODY
        ws.cell(row=row, column=5, value=f'=COUNTIF(\'Master Task List\'!E:E,"{o}")').font = FONT_BODY
        for c in range(4, 6):
            ws.cell(row=row, column=c).border = BORDER_THIN
        row += 1

    # ---- Section C: counts by phase ----
    ws["A14"] = "Tasks by Phase"
    ws["A14"].font = FONT_BODY_BOLD
    _format_header_row(ws, 15, ["Phase", "Title", "Tasks", "Effort (d)", "Blocked", "Done"])

    phases_short = [
        ("P1", "Governance"),
        ("P2", "Architecture"),
        ("P3", "Data/Code Readiness"),
        ("P4", "Platform Foundation"),
        ("P5", "Controls"),
        ("P6", "Code Optimization"),
        ("P7", "Group 1 (2 models)"),
        ("P8a", "Group 2 (8 models)"),
        ("P8b", "Dashboards"),
        ("P9a", "Steady State"),
        ("P9b", "New Model Onboarding"),
        ("M1", "Jenkins: Foundation"),
        ("M2", "Jenkins: AWS-touching"),
        ("M3", "Jenkins: Cutover"),
    ]
    row = 16
    for pid, title in phases_short:
        values = [
            pid, title,
            f'=COUNTIF(\'Master Task List\'!B:B,"{pid}")',
            f'=SUMIF(\'Master Task List\'!B:B,"{pid}",\'Master Task List\'!I:I)',
            f'=COUNTIFS(\'Master Task List\'!B:B,"{pid}",\'Master Task List\'!O:O,"Blocked")',
            f'=COUNTIFS(\'Master Task List\'!B:B,"{pid}",\'Master Task List\'!O:O,"Done")',
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = FONT_BODY_BOLD if pid == "P7" else FONT_BODY
            c.border = BORDER_THIN
        if pid == "P7":
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = FILL_BLOCKED
        row += 1

    _set_col_widths(ws, {"A": 16, "B": 28, "C": 9, "D": 10, "E": 10, "F": 10, "G": 10})


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    out_dir = repo_root / "docs"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "absa-exl-program-plan.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    build_master_task_list(wb)
    build_phase_summary(wb)
    build_milestones(wb)
    build_risk_register(wb)
    build_open_decisions(wb)
    build_dashboard(wb)

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
