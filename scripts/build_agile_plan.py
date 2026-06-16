"""Build the in-depth Agile sprint-plan Excel for the ABSA x EXL delivery team.

Re-anchors the program to the real team start (Sprint 1 = 2026-06-15) and
decomposes it into a capacity-checked EPIC -> Story -> Task backlog across 12
two-week sprints.

Team shape (locked with the user) to make 12 sprints solid without waste:
  8 people. 2 AWS/MLOps engineers split the foundation between Infra and
  Platform/Compute (3 was over-provisioned - the AWS workload is only ~62
  effort-days, well within 2 engineers); 3 SAS developers (two from S1 run
  Group 1 models 1 & 2 in parallel, a third joins S8 for Group 2). Group 2's
  8 models are compressed into S9-S11 so all 10 are live by end S11, leaving
  S12 for hardening + sign-off + handover.

Roles:
  AWS     - AWS/MLOps engineer #1 (Foundation & Infra)
  AWS2    - AWS/MLOps engineer #2 (Platform, Compute & MLOps)
  DE      - Data engineer
  SAS     - SAS developer #1 (Model 1, then Group 2)
  SAS2    - SAS developer #2 (Model 2, then Group 2)
  SAS3    - SAS developer #3 (joins Sprint 8, Group 2)
  DevOps  - DevOps engineer (Jenkins, CI/CD, monitoring)
  TL      - Tech lead / platform engineer (Vishnu)

Sheets:
  1. README              - how to use, capacity model, the scale-up gate, DoR/DoD
  2. Backlog             - EPIC/Story/Task tree (Jira-importable) + points + MoSCoW
  3. Sprint Plan         - dates, goal, milestone, per-role load (formula-driven)
  4. Role Load           - roles x sprints capacity heat grid
  5. Per-Model Tracker   - 10 models x onboarding stages
  6. Governance & Cadence- ceremonies + governance/decision checkpoints
  7. RAID Log            - risks / assumptions / issues / dependencies
  8. Dashboard           - counts by status / role / epic (formula-driven)

Run: `uv run --with openpyxl python scripts/build_agile_plan.py`
Output: docs/absa-exl-agile-plan.xlsx
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------- Styling ----------

FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_BODY_BOLD = Font(name="Arial", size=10, bold=True)
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="1F3864")
FONT_EPIC = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_STORY = Font(name="Arial", size=10, bold=True, color="1F3864")

FILL_HEADER = PatternFill("solid", start_color="1F3864")
FILL_EPIC = PatternFill("solid", start_color="2E4172")
FILL_STORY = PatternFill("solid", start_color="D9E2F3")
FILL_BLOCKED = PatternFill("solid", start_color="FFCDD2")
FILL_MILESTONE = PatternFill("solid", start_color="FFE0B2")
FILL_OK = PatternFill("solid", start_color="C8E6C9")
FILL_WARN = PatternFill("solid", start_color="FFF9C4")
FILL_IDLE = PatternFill("solid", start_color="F2F2F2")
FILL_OFFTEAM = PatternFill("solid", start_color="E0E0E0")

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROLES = ["AWS", "AWS2", "DE", "SAS", "SAS2", "SAS3", "DevOps", "TL"]
ROLE_NAMES = {
    "AWS": "AWS/MLOps Eng #1 - Foundation & Infra",
    "AWS2": "AWS/MLOps Eng #2 - Platform, Compute & MLOps",
    "DE": "Data Engineer",
    "SAS": "SAS Developer #1 (Model 1 / Group 2)",
    "SAS2": "SAS Developer #2 (Model 2 / Group 2)",
    "SAS3": "SAS Developer #3 (Group 2, from S8)",
    "DevOps": "DevOps Engineer",
    "TL": "Tech Lead (Vishnu)",
}
BASE_CAP = {
    "AWS": 8.0, "AWS2": 8.0, "DE": 8.0,
    "SAS": 8.0, "SAS2": 8.0, "SAS3": 8.0, "DevOps": 8.0, "TL": 6.0,
}
SAS3_START_SPRINT = 8  # 3rd SAS dev onboards S8 for Group 2 (confirmed at S5 review)


def capacity(role: str, sprint: int) -> float:
    """Effective effort-day capacity for a role in a given sprint."""
    if role == "SAS3" and sprint < SAS3_START_SPRINT:
        return 0.0
    return BASE_CAP[role]


SPRINT_1_START = date(2026, 6, 15)
N_SPRINTS = 12


def sprint_dates(n: int) -> tuple[date, date]:
    start = SPRINT_1_START + timedelta(days=(n - 1) * 14)
    return start, start + timedelta(days=11)  # Mon -> Fri of week 2


SPRINT_GOALS = {
    1: ("Team onboarded; Jenkins identity decided; python-validate green on Jenkins; ABSA ask-list sent", ""),
    2: ("All no-AWS Jenkins jobs live; AWS bootstrap applied; data contracts started; SAS model-1 review done", ""),
    3: ("Jenkins cutover complete (GHA retired); signing foundation live on real AWS; model-1 optimized", "Jenkins is the sole CI gate"),
    4: ("Registry API live on Lambda (dev); replication path validated; model-1 packaged; threat model done", "Registry live in dev"),
    5: ("SFN executes model-1 in dev; PIR sync; model-2 optimized; CAPACITY REVIEW (scale-up decision)", "*** S5 CAPACITY REVIEW ***"),
    6: ("DRESS REHEARSAL: full chain on real AWS for models 1-2; dashboards + perf harness; POPIA/SARB check", "Dress rehearsal pass"),
    7: ("Group 1 initial production scoring; reconciliation; perf + resilience tests; UAT plan", ""),
    8: ("GROUP 1 ABSA SIGN-OFF; pen-test remediation; 2nd SAS dev onboarded; G2 plan", "*** GROUP 1 SIGN-OFF ***"),
    9: ("Group 2 wave 1 (models 3-5) live; dashboards validated with ABSA; cost dashboard", ""),
    10: ("Group 2 wave 2 (models 6-8) live; output delivery automated", ""),
    11: ("Group 2 wave 3 (models 9-10) live - ALL 10 MODELS SCORING; runbooks; DR verified", "All 10 models live"),
    12: ("HARDENING ONLY (no new models): GROUP 2 SIGN-OFF; handover; hypercare; steady-state go-live", "*** GROUP 2 SIGN-OFF + GO-LIVE ***"),
}

# ---------- EPICs ----------
# (id, title, owner, sprints, objective)
EPICS = [
    ("E01", "Team Onboarding & Governance", "TL", "S1",
     "Everyone productive in week 1: repo green locally, board live, ABSA dependency register running."),
    ("E02", "CI/CD on Existing Jenkins", "DevOps", "S1-S3",
     "Migrate all 6 CI gates from GitHub Actions to the standalone Jenkins per ADR-0011; retire GHA."),
    ("E03", "AWS Foundation (Real Accounts)", "AWS x2", "S1-S4",
     "Bootstrap real EXL accounts, landing zone, signing foundation, registry-on-Lambda, replication path. "
     "Split across 2 AWS engineers: Foundation & Infra (AWS), Platform/Compute/MLOps (AWS2)."),
    ("E04", "Data Engineering & Quality", "DE", "S1-S7",
     "Data contracts, ingestion validation, DQ (volume+drift), reconciliation, lineage, retention, PIR."),
    ("E05", "SAS Model Engineering - Standards & Tooling", "SAS", "S1-S3",
     "SAS prod-readiness standards, real SAS linting, reconciliation framework."),
    ("E06", "ML Compute & Pipeline Execution", "AWS", "S3-S6",
     "Compute platform decision (D04), real Step Functions execution, schedules, prod promote."),
    ("E07", "Security, Controls & Compliance", "TL", "S2-S8",
     "Change workflow, threat model, IAM least-privilege, encryption validation, pen-test, POPIA/SARB, audit pack."),
    ("E08", "Group 1 - Model Onboarding (Models 1-2)", "SAS", "S2-S6",
     "Models 1-2: review, optimize, package, reconcile vs ABSA benchmarks. The proof of the chain."),
    ("E09", "Observability, Dashboards & Delivery", "DevOps", "S4-S9",
     "Run-status dashboards, alerting, DQ/drift panels, cost dashboard, automated signed output delivery."),
    ("E10", "Group 2 - Scale-Out (Models 3-10)", "SAS x3", "S8-S11",
     "Template-driven onboarding across 3 SAS devs: 3 models/sprint in S9-S10, 2 in S11. ALL 10 models "
     "scoring by end S11; S12 is hardening + sign-off only (no new model onboarding)."),
    ("E11", "Steady State, Hypercare & Handover", "TL", "S8-S12",
     "Production readiness, cutover rehearsal, runbooks, DR, hypercare, service review, handover, go-live."),
    ("E12", "Testing, UAT & Performance", "TL", "S5-S8",
     "Integration tests on real AWS, performance/load vs SLA, failure-mode/resilience, UAT with ABSA, regression pack."),
    ("E13", "Capacity & Resourcing", "TL", "S5-S8",
     "2-month capacity review; confirm 3rd SAS dev onboarding (S8) + AWS ramp-down plan after foundation (S7)."),
    ("E14", "Model Implementation Documentation (IDG)", "TL", "S5-S12",
     "LLM-assisted Implementation Document Generator (ADR-0012): drafts the per-version 'as-built' doc from "
     "dev docs + code + platform facts; human-approved; SR 11-7/GMRMG evidence; one living doc per model version."),
]

# ---------- Stories ----------
# (id, epic, title, sprint_label, moscow, acceptance)
STORIES = [
    ("ST-0101", "E01", "Environment & platform onboarding", "S1", "Must",
     "All engineers: 278-test suite green + `make demo` exit 0 locally; walkthrough attended."),
    ("ST-0102", "E01", "Agile working agreement & board", "S1", "Must",
     "Jira board mirrors this backlog; DoR/DoD + ceremony calendar published."),
    ("ST-0103", "E01", "ABSA dependency management", "S1", "Must",
     "8-item ask-list sent; weekly chase cadence running with named SPOC."),
    ("ST-0201", "E02", "Jenkins discovery & identity decision", "S1", "Must",
     "Topology documented; auth path chosen; ADR-0011 amended."),
    ("ST-0202", "E02", "Shared library proving ground", "S1-S2", "Must",
     "absa-ci registered; ci/python-validate status visible on a GitHub PR."),
    ("ST-0203", "E02", "No-AWS pipelines live", "S2", "Must",
     "code-intake, terraform-validate, localstack-demo jobs green on Jenkins."),
    ("ST-0204", "E02", "AWS-touching pipelines + cutover", "S3", "Must",
     "sign/register green vs real AWS; 1-week parallel run byte-identical; branch protection flipped; GHA disabled."),
    ("ST-0301", "E03", "Account bootstrap & state backends", "S1-S2", "Must",
     "CloudTrail/GuardDuty/SecurityHub live in 3 EXL accounts; TF state backends operational."),
    ("ST-0302", "E03", "Landing zone & connectivity", "S2-S3", "Must",
     "Landing zone applied; network path to ABSA implemented; IPs whitelisted."),
    ("ST-0303", "E03", "Signing foundation live", "S2-S4", "Must",
     "KMS CMK in exl-prod; PEM published; cross-account verify passes with ABSA principal."),
    ("ST-0304", "E03", "Registry API on real AWS", "S3-S4", "Must",
     "registry-api on Lambda+APIGW dev->prod; SigV4 register smoke passes."),
    ("ST-0305", "E03", "S3 replication path", "S4", "Must",
     "Replication modules applied; encrypted transfer validated with ABSA test object."),
    ("ST-0401", "E04", "Data contracts per model", "S1-S2", "Must",
     "Data dictionary template + input schemas for models 1-2; volume/cadence matrix for all 10."),
    ("ST-0402", "E04", "Ingestion & arrival validation", "S2-S3", "Must",
     "Bucket layout convention; schema-on-arrival validation; reject/return-to-ABSA workflow."),
    ("ST-0403", "E04", "Data quality framework", "S4-S5", "Must",
     "Volume band checks + PSI drift design (feeds D01); DQ report artifact per run."),
    ("ST-0404", "E04", "PIR integration", "S5-S6", "Must",
     "pir.yaml synced from ABSA PIR system; coverage report on real model inputs."),
    ("ST-0405", "E04", "Group 1 data operations", "S6-S7", "Must",
     "Dress-rehearsal data prepared; DQ tuned on first production runs."),
    ("ST-0406", "E04", "Reconciliation & lineage", "S3-S7", "Should",
     "Source-vs-landing reconciliation; data lineage capture; backfill capability."),
    ("ST-0407", "E04", "Retention implementation", "S7", "Should",
     "S3 retention/archival implemented and verified per model class."),
    ("ST-0501", "E05", "SAS standards & tooling", "S1-S3", "Must",
     "Prod-readiness standards published; SAS runtime obtained; static_sas upgraded to real lint."),
    ("ST-0504", "E05", "Reconciliation framework", "S4-S6", "Must",
     "Diff tool with tolerance bands + per-variable deltas; PIR-ready report template."),
    ("ST-0502", "E08", "Model 1 onboarding (SAS #1)", "S2-S4", "Must",
     "Optimized code reconciles vs ABSA benchmarks; packaged per ADR-0010; code-intake green."),
    ("ST-0503", "E08", "Model 2 onboarding (SAS #2, in parallel)", "S2-S6", "Must",
     "Same exit criteria as model 1; runs in parallel on the 2nd SAS dev to de-risk the proof."),
    ("ST-0601", "E06", "Compute platform decision (D04)", "S3", "Must",
     "ADR records SFN+Lambda vs SageMaker choice with ABSA architecture board."),
    ("ST-0602", "E06", "Step Functions runtime execution", "S5-S6", "Must",
     "Model-1 ASL executes on real AWS in dev; compute layer built; EventBridge schedules set."),
    ("ST-0603", "E06", "Production pipeline infra", "S6", "Must",
     "Prod promote done; SLA bands defined."),
    ("ST-0701", "E07", "Change workflow (D02)", "S5", "Must",
     "CAB/IVU mapped to registry approve route; production-change path documented."),
    ("ST-0702", "E07", "Rotation, retention policy & access audit", "S4-S7", "Must",
     "Rotation cadence doc; S3 lifecycle policies; KMS rotation drill; pre-go-live access audit."),
    ("ST-0703", "E07", "Security assurance", "S4-S8", "Must",
     "Threat model, IAM least-privilege audit, encryption validation, pen-test + remediation."),
    ("ST-0704", "E07", "Compliance evidence", "S6-S8", "Must",
     "POPIA/SARB control verification vs control-matrix; G1 audit evidence pack assembled."),
    ("ST-0801", "E10", "Dress rehearsal (full chain)", "S6", "Must",
     "Producer + verifier chain on real AWS for models 1-2; transcript archived; defects logged."),
    ("ST-0802", "E10", "Group 1 production scoring", "S7", "Must",
     "Initial scheduled runs in prod; outputs delivered; DQ green."),
    ("ST-0803", "E10", "Reconciliation & defect cycle", "S7-S8", "Must",
     "Outputs reconcile within tolerance; defects closed and re-run."),
    ("ST-0804", "E10", "PIR & ABSA sign-off", "S8", "Must",
     "PIR evidence pack submitted; written ABSA Risk Owner sign-off (THE program gate)."),
    ("ST-0901", "E09", "Dashboard tooling & build", "S4-S6", "Must",
     "D08 decided; run-status dashboard live for G1 runs."),
    ("ST-0902", "E09", "Alerting & notifications", "S5-S9", "Must",
     "Failure alerts routed; ABSA SPOC notification path tested."),
    ("ST-0903", "E09", "DQ / drift panels", "S8", "Should",
     "DQ + drift panels visible alongside run status."),
    ("ST-0904", "E09", "Output delivery automation (D09)", "S8-S9", "Must",
     "Signed outputs auto-delivered to ABSA with confirmation evidence; validated with recipients."),
    ("ST-0905", "E09", "Operational monitoring depth", "S6-S9", "Should",
     "Per-pipeline alarms baseline; cost & usage dashboard."),
    ("ST-1001", "E10", "Group 2 onboarding plan", "S8", "Must",
     "Per-model checklist derived from G1 lessons; wave allocation agreed."),
    ("ST-1002", "E10", "Wave 1: models 3-5 (3 SAS devs)", "S9", "Must", "All three models scoring in prod, reconciled, PIR'd."),
    ("ST-1003", "E10", "Wave 2: models 6-8 (3 SAS devs)", "S10", "Must", "Same exit criteria."),
    ("ST-1004", "E10", "Wave 3: models 9-10 - ALL 10 LIVE", "S11", "Must",
     "Final 2 models scoring; all 10 models live by end S11; 3rd SAS dev floats on reconciliation/defects."),
    ("ST-1005", "E10", "S12 hardening + Group 2 sign-off", "S12", "Must",
     "No new models. Final reconciliation sweep, defect closure, KT; written Group 2 sign-off."),
    ("ST-1101", "E11", "Runbooks & support model", "S11", "Must",
     "Incident/recovery/rollback runbooks; support rota + escalation matrix."),
    ("ST-1102", "E11", "Handover & go-live", "S12", "Must",
     "Handover sessions done; service review cadence scheduled; steady-state checklist complete."),
    ("ST-1103", "E11", "Production readiness & hypercare", "S8-S12", "Must",
     "Go/no-go review; cutover rehearsal; DR verified; hypercare defined; ops handbook."),
    ("ST-1201", "E12", "Integration & E2E testing", "S5-S7", "Must",
     "Integration suite on real AWS; perf/load test harness + execution vs SLA; resilience test."),
    ("ST-1202", "E12", "UAT with ABSA", "S7-S8", "Must",
     "UAT plan + scripts; UAT executed with ABSA on Group 1; sign-off captured."),
    ("ST-1203", "E12", "Regression test pack", "S8", "Should",
     "Regression pack for ongoing changes wired into CI."),
    ("ST-1301", "E13", "Capacity review & AWS ramp-down", "S5-S8", "Must",
     "2-month velocity review; confirm 3rd SAS dev onboarding (S8); plan AWS ramp-down/redeploy after S7 foundation."),
    ("ST-1302", "E13", "3rd SAS developer onboarding", "S8", "Must",
     "SAS3 productive: env green, standards studied, G1 reconciliation shadowed."),
    ("ST-1401", "E14", "IDG design + decision", "S5-S6", "Must",
     "ADR-0012 accepted; impl-doc template + section schema; deterministic context-bundle spec; LLM provider DPA confirmed."),
    ("ST-1402", "E14", "IDG build", "S7", "Must",
     "Context bundler (facts from manifest/registry/validation) + LLM provider adapter (Azure OpenAI / Anthropic) + grounded prompt + raw-data/PII guard."),
    ("ST-1403", "E14", "IDG governance + integration", "S7-S8", "Must",
     "Render md+PDF; human review + approval workflow + provenance capture; implementation_doc_ref wired into package manifest + registry; G1 docs generated."),
]

# ---------- Tasks (explicit, Sprints 1-8 IDs stable for the kickoff briefs) ----------
# (id, story, title, role, sprint_n, est_days, deps, blocked_on, notes)
TASKS = [
    # E01
    ("T-0101", "ST-0101", "Clone repo, uv sync, run full 278-test suite locally", "AWS", 1, 0.5, "", "", ""),
    ("T-0102", "ST-0101", "Clone repo, uv sync, run full 278-test suite locally", "DE", 1, 0.5, "", "", ""),
    ("T-0103", "ST-0101", "Clone repo, uv sync, run full 278-test suite locally", "SAS", 1, 0.5, "", "", ""),
    ("T-0104", "ST-0101", "Clone repo, uv sync, run full 278-test suite locally", "DevOps", 1, 0.5, "", "", ""),
    ("T-0105", "ST-0101", "Run `make demo` (LocalStack chain) green locally", "AWS", 1, 0.5, "T-0101", "", ""),
    ("T-0106", "ST-0101", "Run `make demo` (LocalStack chain) green locally", "DE", 1, 0.5, "T-0102", "", ""),
    ("T-0107", "ST-0101", "Run `make demo` (LocalStack chain) green locally", "SAS", 1, 0.5, "T-0103", "", ""),
    ("T-0108", "ST-0101", "Run `make demo` (LocalStack chain) green locally", "DevOps", 1, 0.5, "T-0104", "", ""),
    ("T-0101b", "ST-0101", "Clone repo, uv sync, tests green locally", "AWS2", 1, 0.5, "", "", ""),
    ("T-0103b", "ST-0101", "Clone repo, uv sync, tests green locally", "SAS2", 1, 0.5, "", "", ""),
    ("T-0105b", "ST-0101", "Run `make demo` green locally", "AWS2", 1, 0.5, "T-0101b", "", ""),
    ("T-0107b", "ST-0101", "Run `make demo` green locally", "SAS2", 1, 0.5, "T-0103b", "", ""),
    ("T-0109", "ST-0101", "Host architecture walkthrough: chain-of-custody, repo tour, ADR index", "TL", 1, 1.0, "", "", "Record for future joiners"),
    ("T-0110", "ST-0102", "Stand up Jira/ADO board; import this EPIC/Story/Task tree", "TL", 1, 1.0, "", "", "Backlog sheet is the import source"),
    ("T-0111", "ST-0102", "Publish Definition of Ready/Done + PR / branch-protection workflow", "TL", 1, 0.5, "", "", ""),
    ("T-0112", "ST-0102", "Ceremony calendar: standup, planning, review+demo, retro", "TL", 1, 0.5, "", "", ""),
    ("T-0113", "ST-0103", "Send 8-item dependency ask-list to ABSA SPOC", "TL", 1, 0.5, "", "", "Accounts, IAM, SAS runtime, PIR, data movement, CAB/IVU, network, model docs"),
    ("T-0114", "ST-0103", "Stand up weekly dependency-chase cadence + escalation path", "TL", 1, 0.5, "T-0113", "", ""),
    # E02
    ("T-0201", "ST-0201", "Document Jenkins topology: host, agents, executors, plugins, versions", "DevOps", 1, 1.0, "", "", "Standalone instance already running"),
    ("T-0202", "ST-0201", "Choose AWS auth path: instance profile / OIDC IdP plugin / keys+rotation", "DevOps", 1, 1.0, "T-0201", "", "Drives signing-foundation trust policy"),
    ("T-0203", "ST-0201", "Amend ADR-0011 with the standalone-Jenkins identity decision", "TL", 1, 0.5, "T-0202", "", ""),
    ("T-0204", "ST-0202", "Register absa-ci as Global Pipeline Library", "DevOps", 1, 0.5, "T-0201", "", "ci/jenkins/ path in this repo"),
    ("T-0205", "ST-0202", "Create multibranch job for python-validate; first run", "DevOps", 1, 1.0, "T-0204", "", ""),
    ("T-0206", "ST-0202", "Verify ci/python-validate commit status lands on a GitHub PR", "DevOps", 1, 1.0, "T-0205", "", "Branch-protection sandbox check"),
    ("T-0207", "ST-0202", "Fix shared-library Groovy issues surfaced by first runs", "DevOps", 2, 2.0, "T-0206", "", "publishChecks plugin quirks, PATH handling"),
    ("T-0208", "ST-0203", "code-intake job live", "DevOps", 2, 0.5, "T-0207", "", ""),
    ("T-0209", "ST-0203", "terraform-validate job live (docker tflint/tfsec/checkov/gitleaks)", "DevOps", 2, 1.5, "T-0207", "", ""),
    ("T-0210", "ST-0203", "localstack-demo job live (Docker on agent, 0/1/2/3 exit-code gate)", "DevOps", 2, 2.0, "T-0207", "", "Needs Docker socket on agent"),
    ("T-0211", "ST-0203", "Wire platform secrets into Jenkins credentials store", "DevOps", 2, 1.0, "T-0202", "", "Secrets Manager-backed if reachable"),
    ("T-0212", "ST-0204", "pipeline-factory sign+register stages green against real AWS", "DevOps", 3, 2.0, "T-0309,T-0211", "", ""),
    ("T-0213", "ST-0204", "publish-signing-key job", "DevOps", 3, 1.0, "T-0310", "", ""),
    ("T-0214", "ST-0204", "1-week GHA-vs-Jenkins parallel run; byte-equivalence report", "DevOps", 3, 2.0, "T-0212", "", "Compare manifest digests + signatures"),
    ("T-0215", "ST-0204", "Flip branch protection to Jenkins contexts; GHA -> *.disabled.yml", "DevOps", 3, 0.5, "T-0214", "", ""),
    ("T-0216", "ST-0204", "Docs sweep: ADR-0003/0009 final edits; control-matrix rows -> Live", "TL", 3, 1.0, "T-0215", "", ""),
    # E03
    ("T-0301", "ST-0301", "Module/stack review: gap notes vs real-AWS apply", "AWS", 1, 2.0, "", "", "terraform/ + account-bootstrap"),
    ("T-0302", "ST-0301", "tfvars templates per env (awaiting account IDs)", "AWS", 1, 1.0, "T-0301", "", ""),
    ("T-0303", "ST-0301", "State backend design: S3 + DynamoDB locks per account", "AWS", 1, 1.0, "T-0301", "", ""),
    ("T-0304", "ST-0301", "Apply account-bootstrap to exl-dev/stg/prod", "AWS", 2, 2.0, "T-0302", "ABSA: account IDs", "CloudTrail, GuardDuty, SecurityHub, CIS alarms"),
    ("T-0305", "ST-0301", "Apply state backends", "AWS", 2, 1.0, "T-0304", "", ""),
    ("T-0306", "ST-0302", "Apply landing-zone module (dev first)", "AWS", 2, 2.0, "T-0304", "", ""),
    ("T-0307", "ST-0302", "Implement network connectivity per ABSA decision", "AWS", 3, 3.0, "T-0306", "ABSA: network choice", "VPC peering / TGW / PrivateLink"),
    ("T-0308", "ST-0302", "IP whitelisting coordination with ABSA network team", "AWS", 2, 1.0, "T-0306", "ABSA: network choice", ""),
    ("T-0309", "ST-0303", "identity_provider variable in signing-foundation TF (per T-0202)", "AWS2", 2, 2.0, "T-0202", "", "Supports github_actions | jenkins variants during cutover"),
    ("T-0310", "ST-0303", "Apply signing stack to exl-prod; CMK + buckets + roles live", "AWS2", 3, 1.0, "T-0309,T-0304", "", ""),
    ("T-0311", "ST-0303", "Run publish-key; PEM in public bucket", "AWS2", 3, 0.5, "T-0310", "", ""),
    ("T-0312", "ST-0303", "Cross-account verify test with real ABSA principal", "AWS2", 4, 1.0, "T-0311", "ABSA: IAM ARNs", ""),
    ("T-0313", "ST-0304", "Lambda packaging for registry-api (adapter + artifact build)", "AWS2", 3, 2.0, "T-0304", "", "Largest demo->prod gap per Phase 3 closeout"),
    ("T-0314", "ST-0304", "Finish Lambda packaging + tests", "AWS2", 4, 1.0, "T-0313", "", ""),
    ("T-0315", "ST-0304", "Apply APIGW+Lambda+DynamoDB registry stack (dev) + smoke", "AWS2", 4, 2.0, "T-0314", "", ""),
    ("T-0316", "ST-0304", "SigV4 register smoke from pipeline-factory", "TL", 4, 1.0, "T-0315", "", ""),
    ("T-0317", "ST-0304", "Promote registry to stg + prod", "AWS2", 4, 1.0, "T-0316", "", ""),
    ("T-0318", "ST-0305", "Apply s3-replication source + destination modules", "AWS", 4, 1.5, "T-0306", "ABSA: data movement", ""),
    ("T-0319", "ST-0305", "Encrypted transfer validated end-to-end with ABSA test object", "AWS", 4, 1.0, "T-0318", "ABSA: data movement", ""),
    # E04
    ("T-0401", "ST-0401", "Study platform-contracts schemas + PIR mapping + demo data flow", "DE", 1, 1.0, "", "", ""),
    ("T-0402", "ST-0401", "Data dictionary template (input schema, types, nullability, PIR link)", "DE", 1, 2.0, "T-0401", "", ""),
    ("T-0403", "ST-0401", "Input schemas for models 1-2", "DE", 2, 2.0, "T-0402", "ABSA: model docs", ""),
    ("T-0404", "ST-0401", "Volume/cadence matrix for all 10 models", "DE", 2, 1.0, "T-0402", "", "Daily / weekly / monthly per model"),
    ("T-0405", "ST-0402", "Bucket layout + partitioning convention for scoring inputs", "DE", 2, 1.0, "T-0402", "", ""),
    ("T-0406", "ST-0402", "Arrival validation job: schema check on landing", "DE", 3, 3.0, "T-0405", "", ""),
    ("T-0407", "ST-0402", "Reject / return-to-ABSA workflow for failed files", "DE", 3, 1.0, "T-0406", "", ""),
    ("T-0408", "ST-0403", "Volume checks: row-count expectation bands per model", "DE", 4, 2.0, "T-0406", "", ""),
    ("T-0409", "ST-0403", "Drift metric design (PSI per feature) -> D01 proposal to ABSA", "DE", 4, 2.0, "T-0406", "", "Feeds open decision D01"),
    ("T-0410", "ST-0403", "DQ report artifact published per scoring run", "DE", 5, 2.0, "T-0408", "", ""),
    ("T-0411", "ST-0404", "PIR feed/API contract session with ABSA", "DE", 5, 0.5, "", "ABSA: PIR contract", ""),
    ("T-0412", "ST-0404", "pir.yaml sync from ABSA PIR system", "DE", 5, 3.0, "T-0411", "ABSA: PIR contract", ""),
    ("T-0413", "ST-0404", "PIR coverage report on real model inputs", "DE", 6, 1.0, "T-0412", "", ""),
    ("T-0414", "ST-0405", "Dress-rehearsal data preparation (models 1-2)", "DE", 6, 2.0, "T-0403", "", ""),
    ("T-0415", "ST-0405", "DQ on first production runs + threshold tuning", "DE", 7, 2.0, "T-0410", "", ""),
    ("T-0416", "ST-0406", "Source-vs-landing row-count reconciliation check", "DE", 3, 1.0, "T-0405", "", ""),
    ("T-0417", "ST-0406", "Data lineage capture (input dataset -> run -> output)", "DE", 6, 1.5, "T-0406", "", ""),
    ("T-0418", "ST-0406", "Backfill / re-run capability for missed cadence windows", "DE", 7, 1.5, "T-0406", "", ""),
    ("T-0419", "ST-0407", "Implement + verify S3 retention/archival per model class", "DE", 7, 1.0, "T-0703pol", "", "Pairs with AWS T-0703"),
    # E05
    ("T-0501", "ST-0501", "Study credit-risk-pd package + static_sas checker internals", "SAS", 1, 2.0, "", "", ""),
    ("T-0501b", "ST-0501", "Study credit-risk-pd package + SAS standards (with SAS #1)", "SAS2", 1, 1.5, "", "", ""),
    ("T-0502", "ST-0501", "SAS prod-readiness standards doc (naming, macros, errors, logging)", "SAS", 1, 2.0, "T-0501", "", ""),
    ("T-0503", "ST-0501", "Obtain SAS runtime image + license terms from ABSA", "SAS2", 2, 0.5, "", "ABSA: SAS runtime", ""),
    ("T-0504", "ST-0501", "Upgrade static_sas checker from structural to real lint", "SAS2", 3, 3.0, "T-0503", "ABSA: SAS runtime", "Pairs with TL on checker plumbing"),
    ("T-0514", "ST-0504", "Output diff tool: tolerance bands, per-variable deltas", "SAS", 4, 2.0, "T-0508", "", "Co-design with DE"),
    ("T-0515", "ST-0504", "Reconciliation report template (PIR-ready)", "SAS", 6, 1.0, "T-0514", "", ""),
    # E08 Group 1 model onboarding
    ("T-0505", "ST-0502", "Model-1 dev-code review vs benchmark spec", "SAS", 2, 2.0, "T-0502", "ABSA: model docs", ""),
    ("T-0506", "ST-0502", "Benchmark spec gap notes; clarify with ABSA model owner", "SAS", 2, 1.0, "T-0505", "", ""),
    ("T-0507", "ST-0502", "Optimize + standardize model-1 scoring code (profile, refactor, unit-test)", "SAS", 3, 4.0, "T-0505", "", ""),
    ("T-0508", "ST-0502", "Regression harness vs ABSA benchmark outputs", "SAS", 4, 2.0, "T-0507", "", ""),
    ("T-0509", "ST-0502", "Package model-1 per ADR-0010; code-intake validate green", "SAS", 4, 1.0, "T-0508", "", ""),
    ("T-0510", "ST-0503", "Model-2 dev-code review vs benchmark spec", "SAS2", 3, 2.0, "T-0502", "ABSA: model docs", "Parallel to model-1 on SAS #1"),
    ("T-0511", "ST-0503", "Optimize + standardize model-2 scoring code (profile, refactor, unit-test)", "SAS2", 4, 4.0, "T-0510", "", ""),
    ("T-0512", "ST-0503", "Model-2 regression harness", "SAS2", 5, 2.0, "T-0511", "", ""),
    ("T-0513", "ST-0503", "Package model-2 + code-intake green", "SAS2", 5, 1.0, "T-0512", "", ""),
    ("T-0516", "ST-0502", "Model-1 implementation doc: generate (IDG) + review + approve", "SAS", 8, 0.5, "T-1405", "", "ADR-0012; living doc per version"),
    ("T-0517", "ST-0503", "Model-2 implementation doc: generate (IDG) + review + approve", "SAS2", 8, 0.5, "T-1405", "", "ADR-0012"),
    # E06
    ("T-0601", "ST-0601", "ADR: D04 compute choice (SFN+Lambda vs SageMaker) w/ architecture board", "AWS2", 3, 1.0, "", "", "Gates compute build in S5"),
    ("T-0602", "ST-0602", "Deploy real ASL standard-batch for model-1 (dev)", "AWS2", 5, 3.0, "T-0601,T-0509", "", ""),
    ("T-0603", "ST-0602", "Build compute layer - core (Lambda container / SM Processing per D04)", "AWS2", 5, 2.0, "T-0601", "", ""),
    ("T-0603b", "ST-0602", "Build compute layer - finalize + integrate", "AWS2", 6, 2.0, "T-0603", "", ""),
    ("T-0604", "ST-0602", "EventBridge schedules per model cadence", "AWS2", 6, 1.0, "T-0602", "", ""),
    ("T-0605", "ST-0603", "Promote pipeline infra to prod", "AWS2", 6, 2.0, "T-0602", "", ""),
    ("T-0606", "ST-0603", "Define SLA bands per model cadence", "AWS2", 6, 1.0, "T-0605", "", "Feeds perf test T-1202/1203"),
    # E07
    ("T-0701", "ST-0701", "Map CAB/IVU contract onto registry approve/retire routes (D02)", "TL", 5, 2.0, "", "ABSA: CAB/IVU contract", ""),
    ("T-0702", "ST-0702", "Rotation cadence doc: Jenkins creds, bot PAT, registrar tokens (D03)", "DevOps", 4, 1.0, "T-0211", "", ""),
    ("T-0703pol", "ST-0702", "S3 lifecycle/retention policies authored per retention rules", "AWS", 7, 1.0, "T-0318", "", ""),
    ("T-0704", "ST-0702", "KMS rotation drill per docs/runbooks/kms-key-rotation.md", "DevOps", 7, 1.0, "T-0311", "", "AWS supports 0.5d"),
    ("T-0705", "ST-0702", "Pre-go-live access review + least-privilege audit", "AWS", 7, 1.0, "T-0310", "", ""),
    ("T-0706", "ST-0703", "Threat model workshop (STRIDE) on cross-account signing chain", "TL", 4, 1.5, "", "", "AWS co-runs"),
    ("T-0707", "ST-0703", "IAM least-privilege audit (automated policy analysis)", "AWS2", 7, 1.5, "T-0310", "", ""),
    ("T-0708", "ST-0703", "Encryption validation: in-transit (TLS/replication) + at-rest (KMS) evidence", "AWS2", 7, 1.5, "T-0318", "", ""),
    ("T-0709", "ST-0703", "Pen-test coordination + scope with ABSA security", "DevOps", 7, 1.0, "", "", ""),
    ("T-0710", "ST-0703", "Pen-test remediation window", "DevOps", 8, 1.5, "T-0709", "", ""),
    ("T-0711", "ST-0704", "POPIA/SARB control verification vs control-matrix", "TL", 6, 1.5, "", "", ""),
    ("T-0712", "ST-0704", "Audit evidence pack assembly (G1) for ABSA risk + external audit", "TL", 8, 1.5, "T-0813", "", ""),
    # E10 dress rehearsal + G1 go-live (sign-off)
    ("T-0801", "ST-0801", "Dress rehearsal orchestration: full chain on real AWS, models 1-2", "TL", 6, 2.0, "T-0513,T-0602,T-0317", "", "Producer + verifier, archived transcript"),
    ("T-0802", "ST-0801", "Dress rehearsal infra support + fixes", "AWS", 6, 2.0, "T-0801", "", ""),
    ("T-0803", "ST-0801", "Dress rehearsal model runs + output checks", "SAS", 6, 2.0, "T-0801", "", ""),
    ("T-0804", "ST-0801", "Dress rehearsal CI/pipeline ops", "DevOps", 6, 1.0, "T-0801", "", ""),
    ("T-0805", "ST-0802", "Initial production scoring runs (models 1-2)", "AWS2", 7, 2.0, "T-0801", "", ""),
    ("T-0806", "ST-0802", "Production run monitoring + incident handling", "DevOps", 7, 1.0, "T-0805", "", ""),
    ("T-0807", "ST-0803", "Reconcile G1 outputs vs benchmarks (tolerance report)", "SAS", 7, 3.0, "T-0805,T-0514", "", ""),
    ("T-0808", "ST-0803", "Defect triage + fix cycle", "SAS2", 7, 3.0, "T-0807", "", ""),
    ("T-0809", "ST-0803", "Defect triage facilitation + ABSA comms", "TL", 7, 1.0, "T-0807", "", ""),
    ("T-0810", "ST-0803", "Defect closure + re-run + final reconciliation", "SAS", 8, 2.0, "T-0808", "", ""),
    ("T-0811", "ST-0804", "PIR evidence pack (reconciliation, DQ, audit trail)", "SAS", 8, 2.0, "T-0810", "", ""),
    ("T-0812", "ST-0804", "PIR data evidence: lineage + DQ reports", "DE", 8, 1.0, "T-0810", "", ""),
    ("T-0813", "ST-0804", "ABSA sign-off review meeting + written sign-off", "TL", 8, 2.0, "T-0811", "", "*** THE program gate ***"),
    ("T-0814", "ST-0804", "G1 production schedules confirmed live", "AWS2", 8, 0.5, "T-0813", "", ""),
    # E09
    ("T-0901", "ST-0901", "D08 tooling decision: CloudWatch / Grafana / QuickSight", "DevOps", 4, 1.5, "", "", ""),
    ("T-0902", "ST-0901", "Run-status dashboard build (runs, exceptions, durations)", "DevOps", 6, 3.0, "T-0901", "", ""),
    ("T-0903", "ST-0902", "Alert routing: SNS -> email/Slack; severity mapping", "DevOps", 5, 2.0, "T-0901", "", ""),
    ("T-0904", "ST-0902", "Failure notifications to ABSA SPOC path tested", "DevOps", 9, 1.0, "T-0903", "", ""),
    ("T-0905", "ST-0903", "DQ + drift panels alongside run status", "DE", 8, 2.0, "T-0410,T-0902", "", ""),
    ("T-0906", "ST-0904", "Automated signed output delivery to ABSA + confirmation evidence", "DevOps", 8, 2.0, "T-0319", "ABSA: data movement", ""),
    ("T-0907", "ST-0904", "Delivery validation with ABSA recipients", "DevOps", 9, 1.0, "T-0906", "", ""),
    ("T-0908", "ST-0904", "Dashboard access + validation with ABSA recipients", "DevOps", 9, 1.0, "T-0902", "", ""),
    ("T-0909", "ST-0905", "Per-pipeline CloudWatch alarms + dashboard baseline", "DevOps", 6, 1.0, "T-0901", "", ""),
    ("T-0910", "ST-0905", "Cost & usage dashboard", "DevOps", 9, 1.0, "T-0901", "", ""),
    # E10 Group 2 plan + sign-off (per-model tasks generated below)
    ("T-1001", "ST-1001", "Group 2 onboarding plan + per-model checklist from G1 lessons", "TL", 8, 1.0, "T-0813", "", ""),
    ("T-1025", "ST-1005", "Group 2 sign-off pack + ABSA meeting", "TL", 12, 2.0, "", "", "Closes initial onboarding scope"),
    # E11
    ("T-1101", "ST-1101", "Runbooks: incident, recovery, rollback per model family", "DevOps", 11, 3.0, "T-0806", "", ""),
    ("T-1102", "ST-1101", "Support rota + escalation matrix", "TL", 11, 1.0, "T-1101", "", ""),
    ("T-1103", "ST-1101", "Cost review + right-sizing pass", "AWS2", 11, 2.0, "T-0814", "", ""),
    ("T-1104", "ST-1102", "Quarterly service review template + first review scheduled", "TL", 11, 1.0, "T-1102", "", ""),
    ("T-1105", "ST-1102", "Handover sessions: ops team + ABSA", "TL", 12, 2.0, "T-1101", "", ""),
    ("T-1106", "ST-1102", "Handover support + docs finalization", "DevOps", 12, 1.0, "T-1105", "", ""),
    ("T-1107", "ST-1102", "Steady-state go-live checklist + cutover", "TL", 12, 1.0, "T-1105", "", "Go-live ~Dec 1, 2026"),
    ("T-1108", "ST-1103", "G1 production readiness review (go/no-go checklist)", "DevOps", 8, 0.5, "T-0810", "", ""),
    ("T-1109", "ST-1103", "Cutover rehearsal (G1)", "DevOps", 8, 1.0, "T-1108", "", ""),
    ("T-1110", "ST-1103", "DR plan + backup/restore verification", "AWS2", 11, 1.5, "T-0814", "", ""),
    ("T-1111", "ST-1103", "Define hypercare period + staffing", "TL", 12, 0.5, "T-1107", "", ""),
    ("T-1112", "ST-1103", "Ops handbook / knowledge base", "DevOps", 12, 1.0, "T-1101", "", ""),
    # E12 Testing/UAT/Perf
    ("T-1201", "ST-1201", "Integration test suite on real AWS (beyond LocalStack)", "TL", 5, 1.5, "T-0317", "", ""),
    ("T-1202", "ST-1201", "Performance/load test harness", "DevOps", 6, 1.5, "T-0605", "", ""),
    ("T-1203", "ST-1201", "Performance test execution vs SLA bands", "DevOps", 7, 1.5, "T-1202,T-0606", "", ""),
    ("T-1204", "ST-1201", "Failure-mode / resilience test (retry, partial failure, teardown)", "DevOps", 7, 1.0, "T-0805", "", ""),
    ("T-1205", "ST-1202", "UAT plan + scripts with ABSA", "TL", 7, 1.5, "", "", ""),
    ("T-1206", "ST-1202", "UAT execution support (G1) with ABSA", "DE", 8, 1.5, "T-1205", "", ""),
    ("T-1207", "ST-1202", "UAT sign-off capture", "TL", 8, 0.5, "T-1206", "", ""),
    ("T-1208", "ST-1203", "Regression test pack for ongoing changes", "DevOps", 8, 1.0, "T-0214", "", ""),
    # E13 Capacity & resourcing
    ("T-1301", "ST-1301", "2-month capacity & velocity review; confirm 3rd SAS dev + AWS ramp-down plan", "TL", 5, 0.5, "", "", "*** S5 CAPACITY REVIEW ***"),
    ("T-1302", "ST-1301", "Initiate 3rd SAS dev sourcing/transfer; plan AWS redeploy after S7", "TL", 6, 0.5, "T-1301", "", ""),
    ("T-1303", "ST-1301", "Onboard 3rd SAS dev: access + repo + walkthrough", "TL", 8, 0.5, "T-1302", "", ""),
    ("T-1304", "ST-1302", "Clone repo, uv sync, tests green", "SAS3", 8, 0.5, "T-1303", "", ""),
    ("T-1305", "ST-1302", "Run `make demo` green locally", "SAS3", 8, 0.5, "T-1304", "", ""),
    ("T-1306", "ST-1302", "Study SAS standards + a packaged G1 model", "SAS3", 8, 1.0, "T-1304", "", ""),
    ("T-1307", "ST-1302", "Shadow a G1 reconciliation cycle", "SAS3", 8, 1.0, "T-0810", "", ""),
    # E14 - Implementation Document Generator (ADR-0012)
    ("T-1407", "ST-1401", "ADR-0012 + confirm LLM provider DPA (Azure OpenAI / Anthropic) with ABSA", "TL", 5, 0.5, "", "ABSA: LLM data-processing terms", "*** gates first real IDG run ***"),
    ("T-1401", "ST-1401", "Design impl-doc template + section schema + deterministic context-bundle spec", "TL", 6, 1.5, "T-1407", "", "Co-design with SAS"),
    ("T-1402", "ST-1402", "Build context bundler (facts from manifest/registry/validation summary)", "AWS2", 7, 2.0, "T-1401,T-0317", "", "Facts injected verbatim, never LLM-authored"),
    ("T-1403", "ST-1402", "LLM provider adapter (Azure OpenAI / Anthropic) + grounded prompt", "TL", 7, 2.0, "T-1401", "ABSA: LLM data-processing terms", "Provider-swappable behind an adapter"),
    ("T-1404", "ST-1402", "Raw-data / PII guard on the context bundle (pre-flight checker)", "DE", 7, 1.0, "T-1402", "", "Bundle = code+docs+metadata only; fail on data payload"),
    ("T-1405", "ST-1403", "Render md + PDF; human review + approval workflow + provenance capture", "AWS2", 8, 2.0, "T-1403", "", ""),
    ("T-1406", "ST-1403", "Wire implementation_doc_ref into package-manifest + registry schema (regen models)", "DE", 8, 1.5, "T-1404", "", "Digest-reference pattern like python_pyproject_ref"),
]

# ---------- Group 2 per-model tasks (generated) ----------
# 3 SAS devs run 3 models/sprint in S9-S10 and the final 2 in S11.
# ALL 10 models live by end S11; S12 is hardening + sign-off only.
# model -> (sprint, sas_role, wave_story)
G2_MODELS = [
    (3, 9, "SAS", "ST-1002"),
    (4, 9, "SAS2", "ST-1002"),
    (5, 9, "SAS3", "ST-1002"),
    (6, 10, "SAS", "ST-1003"),
    (7, 10, "SAS2", "ST-1003"),
    (8, 10, "SAS3", "ST-1003"),
    (9, 11, "SAS", "ST-1004"),
    (10, 11, "SAS2", "ST-1004"),
]
# (suffix, title, role_key, est)  role_key "SASX" -> the model's sas owner
G2_TEMPLATE = [
    ("DC", "contract + input schema", "DE", 1.0),
    ("DQ", "DQ rules + verification", "DE", 0.5),
    ("RV", "code review vs benchmark spec", "SASX", 1.0),
    ("OPT", "profile + optimize + unit-test scoring code", "SASX", 2.5),
    ("PKG", "package per ADR-0010 + code-intake green", "SASX", 0.5),
    ("REC", "reconcile vs benchmark + defect fix", "SASX", 1.5),
    ("PIR", "PIR pack + sign-off prep", "SASX", 0.5),
    ("IDD", "implementation doc: generate (IDG) + approve", "SASX", 0.5),
    ("PIPE", "pipeline config + EventBridge schedule", "AWS", 0.5),
    ("DEP", "deploy to prod + smoke", "AWS", 0.5),
    ("MON", "CI + monitoring wiring", "DevOps", 0.5),
]
for _m, _sprint, _sas, _story in G2_MODELS:
    for _suf, _title, _rk, _est in G2_TEMPLATE:
        _role = _sas if _rk == "SASX" else _rk
        TASKS.append(
            (f"T-M{_m:02d}-{_suf}", _story, f"Model {_m}: {_title}", _role, _sprint, _est,
             "T-1001", "", f"Group 2; SAS owner {_sas}")
        )
# S11: 3rd SAS dev floats on reconciliation depth + defect cleanup across the wave
TASKS.append(
    ("T-M-FLOAT", "ST-1004", "Reconciliation depth + defect sweep across wave 3", "SAS3", 11, 4.0,
     "T-M09-OPT,T-M10-OPT", "", "Spare SAS capacity = defect buffer")
)
# Per-wave TL governance checkpoint
for _wave_sprint, _wave_story in [(9, "ST-1002"), (10, "ST-1003"), (11, "ST-1004")]:
    TASKS.append(
        (f"T-W{_wave_sprint}", _wave_story, f"Wave review + ABSA checkpoint (S{_wave_sprint})", "TL",
         _wave_sprint, 0.5, "", "", "")
    )
# S12 hardening (NO new models) - the slack the 9-person team buys
TASKS += [
    ("T-1501", "ST-1005", "Final cross-model reconciliation sweep + tolerance sign-off", "SAS", 12, 3.0, "", "", ""),
    ("T-1502", "ST-1005", "Group 2 defect closure + re-run verification", "SAS2", 12, 3.0, "", "", ""),
    ("T-1503", "ST-1005", "Group 2 knowledge transfer + per-model runbooks", "SAS3", 12, 2.0, "", "", ""),
    ("T-1504", "ST-1005", "Group 2 DQ + lineage evidence pack", "DE", 12, 1.5, "", "", ""),
]

# ---------- Per-Model Tracker data ----------
# (model, group, sprints, sas_owner, cadence)
MODELS = [
    ("Model 1", "Group 1", "S2-S8", "SAS", "TBD (ABSA)"),
    ("Model 2", "Group 1", "S2-S8", "SAS2", "TBD (ABSA)"),
    ("Model 3", "Group 2 W1", "S9", "SAS", "TBD"),
    ("Model 4", "Group 2 W1", "S9", "SAS2", "TBD"),
    ("Model 5", "Group 2 W1", "S9", "SAS3", "TBD"),
    ("Model 6", "Group 2 W2", "S10", "SAS", "TBD"),
    ("Model 7", "Group 2 W2", "S10", "SAS2", "TBD"),
    ("Model 8", "Group 2 W2", "S10", "SAS3", "TBD"),
    ("Model 9", "Group 2 W3", "S11", "SAS", "TBD"),
    ("Model 10", "Group 2 W3", "S11", "SAS2", "TBD"),
]
MODEL_STAGES = ["Contract", "Review", "Optimize", "Package", "Deploy", "Reconcile", "Impl doc", "PIR", "Sign-off"]

# ---------- Governance & cadence ----------
GOV_CEREMONIES = [
    ("Daily standup", "Every working day, 15 min", "All", "Blockers surfaced same-day"),
    ("Sprint planning", "1st Monday of each sprint, 2h", "All", "Sprint backlog committed"),
    ("Backlog refinement", "Mid-sprint Wednesday, 1h", "TL + leads", "Next sprint stories Ready"),
    ("Sprint review + demo", "2nd Friday, 1h", "All + ABSA optional", "Working increment shown"),
    ("Retrospective", "2nd Friday, 45 min", "All", "Actions logged + owned"),
]
# (checkpoint, sprint, date_sprint_n, owner, gate)
GOV_CHECKPOINTS = [
    ("Kickoff + SPOCs confirmed", 1, "TL", "Charter + SPOCs signed"),
    ("Jenkins cutover (GHA retired)", 3, "DevOps", "Jenkins sole CI gate; branch protection flipped"),
    ("Registry live in dev", 4, "AWS", "SigV4 register smoke green"),
    ("S5 CAPACITY REVIEW (confirm 3rd SAS + AWS ramp-down)", 5, "TL", "Confirm SAS3 onboarding S8; plan AWS redeploy after S7"),
    ("Dress rehearsal pass", 6, "TL", "Full chain on real AWS, models 1-2"),
    ("G1 production readiness (go/no-go)", 8, "DevOps", "Readiness checklist + cutover rehearsal clean"),
    ("GROUP 1 ABSA SIGN-OFF", 8, "ABSA Risk", "Written sign-off; PIR + audit pack accepted"),
    ("Steering checkpoint (mid-program)", 6, "TL", "Status + risk review with ABSA + EXL sponsors"),
    ("ALL 10 MODELS LIVE (Group 2 onboarding complete)", 11, "TL", "Models 3-10 scoring; S12 reserved for hardening"),
    ("Steering checkpoint (pre-go-live)", 11, "TL", "Go-live readiness review"),
    ("GROUP 2 SIGN-OFF + go-live", 12, "ABSA Risk", "All 10 models signed off; steady-state begins"),
]

# ---------- RAID ----------
# (type, id, item, owner, sprint/impact, status/mitigation)
RAID = [
    ("Dependency", "DEP-01", "AWS account IDs (3 EXL + ABSA receiving)", "ABSA", "T-0304 (S2)", "Open - bootstrap blocked; slips downstream day-for-day"),
    ("Dependency", "DEP-02", "IAM principal ARNs (kms:Verify / s3:GetObject)", "ABSA", "T-0312 (S4)", "Open - cross-account verify untestable"),
    ("Dependency", "DEP-03", "SAS runtime Docker image + license", "ABSA", "T-0503/0504 (S2-S3)", "Open - SAS validation stays structural-only"),
    ("Dependency", "DEP-04", "PIR system API/feed contract", "ABSA", "T-0411 (S5)", "Open - pir.yaml hand-maintained"),
    ("Dependency", "DEP-05", "Data movement decision (S3 replication vs SFTP)", "ABSA", "T-0318/0906 (S4/S8)", "Open - transfer + delivery blocked"),
    ("Dependency", "DEP-06", "CAB / IVU API contract", "ABSA", "T-0701 (S5)", "Open - change workflow manual"),
    ("Dependency", "DEP-07", "Network connectivity choice (peering/TGW/PrivateLink)", "ABSA", "T-0307 (S3)", "Open - data plane blocked"),
    ("Dependency", "DEP-08", "Approved model docs + code + benchmarks (1-2 first)", "ABSA", "T-0403/0505 (S2)", "Open - SAS + DE model work blocked"),
    ("Dependency", "DEP-09", "LLM data-processing terms (Azure OpenAI / Anthropic): no-retention DPA + region", "ABSA", "T-1407/1403 (S5/S7)", "Open - gates first real IDG run (ADR-0012); manual impl-doc fallback otherwise"),
    ("Risk", "RISK-01", "ABSA account onboarding delayed past S2", "TL", "High / Med", "2 AWS engineers + ~2d/sprint slack in S2-S5 absorb a slip; pre-stage TF + tfvars; LocalStack keeps CI green; escalate end S1"),
    ("Risk", "RISK-02", "Group 1 sign-off (S8) slips, cascades to Group 2", "ABSA Risk", "Med / High", "Models 1-2 run in PARALLEL (2 SAS) from S2; dress rehearsal S6; defect buffer S7-S8"),
    ("Risk", "RISK-03", "Compute platform choice (D04) blocks S5 build", "TL", "Med / Med", "Decision required end S3 at architecture board; AWS2 owns it"),
    ("Risk", "RISK-04", "Benchmark reconciliation discrepancies force rewrite", "SAS", "Med / High", "Incremental validation; per-variable delta report; tolerance bands agreed early; S11 float + S12 hardening absorb rework"),
    ("Risk", "RISK-05", "3rd SAS dev not sourced in time for S8 onboarding", "TL", "Med / Med", "Start sourcing S6; 2 SAS from S1 already cover Group 2 at ~2 models/sprint as fallback (S9-S12)"),
    ("Risk", "RISK-06", "Signature byte-equivalence fails GHA vs Jenkins", "DevOps", "Low / High", "S3 parallel-run week compares digests before cutover"),
    ("Risk", "RISK-07", "DevOps single-person peaks (S2 7.0, S6 6.5) - no DevOps redundancy", "TL", "Med / Med", "TL pairs on Jenkins/dashboards at peaks; only un-doubled role - watch at S2/S6"),
    ("Assumption", "ASM-01", "9-person team funded from S1 (3 AWS, 3 SAS [1 from S8], DE, DevOps, TL)", "TL", "All", "Max-coverage shape chosen to make 12 sprints solid with absorption buffer"),
    ("Assumption", "ASM-02", "AWS2 redeploys or ramps down after S7 once foundation is built", "TL", "S8+", "Foundation is front-loaded; AWS load drops sharply S8-S12. 3rd AWS dropped (was under-utilized) - 2 carry the ~62-day AWS workload"),
    ("Assumption", "ASM-03", "All 10 models fit standard-batch / scalable-batch templates", "TL", "S6+", "Realtime tier is a placeholder; a realtime model would add scope"),
    ("Assumption", "ASM-04", "ABSA benchmarks are deterministic + reproducible", "SAS", "S4+", "Required for tolerance-band reconciliation to be meaningful"),
    ("Assumption", "ASM-05", "IDG sends the LLM code + docs + metadata ONLY - never raw data / PII", "TL", "S7+", "ADR-0012 data-minimisation guard; keeps IDG consistent with the data-residency posture (ADR-0001)"),
    ("Issue", "ISS-00", "(none logged yet - populate during delivery)", "TL", "-", "-"),
]

PREREQS_INTERNAL = [
    ("PRE-01", "Jira/ADO project + board created", "TL", "Before S1 planning"),
    ("PRE-02", "GitHub repo write access for all engineers + CODEOWNERS updated", "TL", "Before S1 day 1"),
    ("PRE-03", "Jenkins admin access for DevOps engineer", "DevOps", "Before S1 day 1"),
    ("PRE-04", "AWS sandbox account access for AWS engineer", "AWS", "Before S1 day 1"),
    ("PRE-05", "Dev machines: uv, Docker Desktop, Terraform 1.9.5, Python 3.12", "All", "S1 day 1-2"),
    ("PRE-06", "Slack/Teams channel + ceremony invites", "TL", "Before S1 planning"),
    ("PRE-07", "SAS desktop/dev license (interim until ABSA runtime)", "SAS", "S1"),
]


# ---------- Validation ----------

def validate_capacity() -> list[str]:
    load: dict[tuple[str, int], float] = {}
    for t in TASKS:
        key = (t[3], t[4])
        load[key] = load.get(key, 0.0) + t[5]
    warnings = []
    for (role, sprint), days in sorted(load.items(), key=lambda x: (x[0][1], x[0][0])):
        cap = capacity(role, sprint)
        if days > cap + 1e-9:
            warnings.append(f"OVERLOAD {role} S{sprint}: {days:.1f}d > {cap:.1f}d cap")
    return warnings


def story_points(days: float) -> int:
    for threshold, pts in [(1, 1), (2, 2), (3, 3), (5, 5), (8, 8)]:
        if days <= threshold:
            return pts
    return 13


# ---------- Sheet builders ----------

def _header(ws, row, cols):
    for i, label in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN


def _widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_readme(wb):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ABSA x EXL - Agile Sprint Plan (in-depth)"
    ws["A1"].font = FONT_TITLE
    lines = [
        "",
        "Anchoring: Sprint 1 = Mon 2026-06-15. 12 sprints x 2 weeks ~ 6 months.",
        "Group 1 ABSA sign-off targeted end of Sprint 8 (2026-10-02). All 10 models live by end S11; Group 2 sign-off + go-live end S12.",
        "",
        "TEAM SHAPE - 8 people (right-sized for a solid 12 sprints):",
        "  - 2 AWS/MLOps engineers split the foundation: AWS (Foundation & Infra) + AWS2 (Platform, Compute &",
        "    MLOps). The total AWS workload is only ~62 effort-days, so 2 carry it at a healthy ~75% peak;",
        "    a 3rd was dropped as under-utilized (it sat idle most of S8-S12).",
        "  - 3 SAS developers: two from S1 run Group 1 models 1 & 2 IN PARALLEL (kills bus-factor-of-one on",
        "    the proof); a third onboards S8 for Group 2.",
        "  Group 2 (8 models) is compressed into S9-S11 (3 models/sprint, then 2) so ALL 10 are live by end S11,",
        "  leaving S12 entirely for hardening, sign-off, and handover - not a race to finish models.",
        "  The S5 capacity review (T-1301) confirms the 3rd SAS dev + plans AWS ramp-down after foundation.",
        "",
        "Team & capacity (effort-days per 2-week sprint = 10 working days x 0.8 focus factor):",
        "  AWS    - AWS/MLOps Eng #1 (Foundation & Infra)        - 8.0 d/sprint",
        "  AWS2   - AWS/MLOps Eng #2 (Platform, Compute & MLOps) - 8.0 d/sprint",
        "  DE     - Data Engineer                                - 8.0 d/sprint",
        "  SAS    - SAS Developer #1 (Model 1 / Group 2)         - 8.0 d/sprint",
        "  SAS2   - SAS Developer #2 (Model 2 / Group 2)         - 8.0 d/sprint",
        "  SAS3   - SAS Developer #3 (Group 2)                   - 8.0 d/sprint FROM SPRINT 8 (0 before)",
        "  DevOps - DevOps Engineer                              - 8.0 d/sprint",
        "  TL     - Tech Lead (Vishnu)                           - 6.0 d/sprint (splits with program mgmt)",
        "",
        "  COST LEVER: AWS2 is loaded S1-S7 (platform + compute) then light. After S7 it can ramp down or",
        "  redeploy to MLOps/optimization work - planned at the S5 review (RAID ASM-02). AWS #1 carries the",
        "  Group 2 pipeline-deploy tail S9-S11.",
        "  Only DevOps is un-doubled; watch its S2 (7.0) and S6 (6.5) peaks (RAID RISK-07).",
        "",
        "How to use:",
        "  1. 'Backlog' is the source of truth: EPIC -> Story -> Task with role, sprint, estimate, points, MoSCoW.",
        "     Import into Jira (CSV) or mirror manually; Sprint 1-8 base task IDs are stable.",
        "  2. 'Sprint Plan' = goal + milestone per sprint + per-role load (formula-driven), 9 role columns.",
        "  3. 'Role Load' = capacity heat grid. Grey = off-team (SAS3 pre-S8). Red = overloaded (must re-plan).",
        "  4. 'Per-Model Tracker' = the 10 models x onboarding stages; the heart of the program.",
        "  5. 'Governance & Cadence' = ceremonies + decision/sign-off checkpoints with target sprints.",
        "  6. 'RAID Log' = risks, assumptions, issues, dependencies (incl. the 8 ABSA dependencies).",
        "  7. Update Status on Backlog as work moves; Dashboard recalculates.",
        "",
        "Definition of Ready (a story may enter a sprint when):",
        "  - Acceptance criteria written; dependencies identified; estimate agreed; no unresolved ABSA blocker for that sprint.",
        "Definition of Done (a task is done when):",
        "  - Merged to main; CI green; docs updated if behaviour changed; acceptance criteria met; demoed on review Friday.",
        "",
        "Context: the platform (registry, pipeline-factory, signer, code-intake, LocalStack demo) is already built and",
        "regression-tested - see docs/phase-3-closeout.md. This plan covers real-AWS deployment, the Jenkins CI",
        "migration, onboarding 10 SAS models, security/compliance hardening, testing/UAT, and go-live.",
    ]
    for i, line in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=line)
        c.font = FONT_BODY
        if line.startswith("CAPACITY") or line.startswith("Definition") or line.startswith("How to use") or line.startswith("Team &"):
            c.font = FONT_BODY_BOLD
    ws.column_dimensions["A"].width = 120


def build_backlog(wb):
    ws = wb.create_sheet("Backlog")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Backlog - EPIC / Story / Task"
    ws["A1"].font = FONT_TITLE

    cols = ["Type", "ID", "Parent", "Epic", "Title", "Role", "Sprint", "Est (d)", "Pts",
            "MoSCoW", "Depends On", "Blocked On (ABSA)", "Status", "Acceptance / Notes"]
    _header(ws, 3, cols)

    stories_by_epic: dict[str, list] = {}
    for s in STORIES:
        stories_by_epic.setdefault(s[1], []).append(s)
    tasks_by_story: dict[str, list] = {}
    for t in TASKS:
        tasks_by_story.setdefault(t[1], []).append(t)

    row = 4
    for epic in EPICS:
        eid, etitle, eowner, esprints, eobj = epic
        vals = ["EPIC", eid, "", eid, f"{etitle} - {eobj}", eowner, esprints, "", "", "", "", "", "", ""]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = FONT_EPIC
            c.fill = FILL_EPIC
            c.alignment = ALIGN_LEFT
            c.border = BORDER_THIN
        row += 1
        for st in stories_by_epic.get(eid, []):
            sid, _, stitle, ssprint, smoscow, sacc = st
            story_days = sum(t[5] for t in tasks_by_story.get(sid, []))
            vals = ["Story", sid, eid, eid, stitle, "", ssprint, story_days,
                    story_points(story_days), smoscow, "", "", "", sacc]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = FONT_STORY
                c.fill = FILL_STORY
                c.alignment = ALIGN_LEFT
                c.border = BORDER_THIN
            row += 1
            for t in tasks_by_story.get(sid, []):
                tid, _, ttitle, trole, tsprint, test_d, tdeps, tblocked, tnotes = t
                vals = ["Task", tid, sid, eid, ttitle, trole, f"S{tsprint}", test_d, "",
                        "", tdeps, tblocked, "Planned", tnotes]
                for ci, v in enumerate(vals, start=1):
                    c = ws.cell(row=row, column=ci, value=v)
                    c.font = FONT_BODY
                    c.alignment = ALIGN_LEFT
                    c.border = BORDER_THIN
                if tblocked:
                    ws.cell(row=row, column=12).fill = FILL_BLOCKED
                row += 1

    _widths(ws, {"A": 7, "B": 11, "C": 9, "D": 6, "E": 58, "F": 8, "G": 8, "H": 7, "I": 5,
                 "J": 9, "K": 14, "L": 18, "M": 10, "N": 44})
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:N{row - 1}"


def build_sprint_plan(wb):
    ws = wb.create_sheet("Sprint Plan")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Sprint Plan - 12 x 2-week sprints"
    ws["A1"].font = FONT_TITLE

    cols = ["Sprint", "Start", "End", "Goal", "Milestone"] + [f"{r} (d)" for r in ROLES] + ["Total (d)"]
    _header(ws, 3, cols)

    row = 4
    for n in range(1, N_SPRINTS + 1):
        s, e = sprint_dates(n)
        goal, milestone = SPRINT_GOALS[n]
        ws.cell(row=row, column=1, value=f"S{n}").font = FONT_BODY_BOLD
        ws.cell(row=row, column=2, value=s).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=3, value=e).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=4, value=goal).font = FONT_BODY
        mc = ws.cell(row=row, column=5, value=milestone)
        mc.font = FONT_BODY_BOLD
        if milestone:
            for ci in range(1, 13):
                ws.cell(row=row, column=ci).fill = FILL_MILESTONE
        for ri, role in enumerate(ROLES):
            col = 6 + ri
            ws.cell(row=row, column=col,
                    value=f'=SUMIFS(Backlog!$H:$H,Backlog!$F:$F,"{role}",Backlog!$G:$G,"S{n}",Backlog!$A:$A,"Task")')
        total_col = 6 + len(ROLES)
        ws.cell(row=row, column=total_col,
                value=f"=SUM({_col_letter(6)}{row}:{_col_letter(total_col - 1)}{row})").font = FONT_BODY_BOLD
        for ci in range(1, total_col + 1):
            c = ws.cell(row=row, column=ci)
            c.border = BORDER_THIN
            c.alignment = ALIGN_LEFT
        row += 1

    _widths(ws, {"A": 7, "B": 11, "C": 11, "D": 66, "E": 26,
                 "F": 8, "G": 8, "H": 8, "I": 8, "J": 9, "K": 8, "L": 9})
    ws.freeze_panes = "A4"


def _col_letter(idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def build_role_load(wb):
    ws = wb.create_sheet("Role Load")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Role Load vs Capacity (effort-days per sprint)"
    ws["A1"].font = FONT_TITLE

    cols = ["Role"] + [f"S{n}" for n in range(1, N_SPRINTS + 1)]
    _header(ws, 3, cols)

    # Precompute load
    load: dict[tuple[str, int], float] = {}
    for t in TASKS:
        load[(t[3], t[4])] = load.get((t[3], t[4]), 0.0) + t[5]

    row = 4
    for role in ROLES:
        ws.cell(row=row, column=1, value=f"{role} - {ROLE_NAMES[role]}").font = FONT_BODY_BOLD
        ws.cell(row=row, column=1).border = BORDER_THIN
        for n in range(1, N_SPRINTS + 1):
            col = 1 + n
            days = load.get((role, n), 0.0)
            cap = capacity(role, n)
            c = ws.cell(row=row, column=col,
                        value=f'=SUMIFS(Backlog!$H:$H,Backlog!$F:$F,"{role}",Backlog!$G:$G,"S{n}",Backlog!$A:$A,"Task")')
            c.alignment = ALIGN_CENTER
            c.border = BORDER_THIN
            # Build-time fill reflecting the committed plan
            if cap == 0.0:
                c.fill = FILL_OFFTEAM
            elif days > cap + 1e-9:
                c.fill = FILL_BLOCKED
            elif days >= cap - 1.0 and days > 0:
                c.fill = FILL_WARN
            elif days > 0:
                c.fill = FILL_OK
            else:
                c.fill = FILL_IDLE
        row += 1

    # Capacity reference row
    ws.cell(row=row, column=1, value="Capacity (cap)").font = FONT_BODY_BOLD
    for n in range(1, N_SPRINTS + 1):
        caps = {capacity(r, n) for r in ROLES}
        # show TL/others note: just show 8 (6 TL); simplest is per-sprint max engineer cap
        ws.cell(row=row, column=1 + n, value="8 / TL 6").alignment = ALIGN_CENTER
        ws.cell(row=row, column=1 + n).font = Font(name="Arial", size=8, color="808080")
    row += 2
    legend = ("Green = loaded within capacity | Yellow = near cap (<=1d slack) | "
              "Red = over capacity (re-plan) | Grey = off-team (SAS3 before S8) | Light = idle. "
              "Fills reflect the committed plan; values recalc live via SUMIFS. "
              "AWS2 idle late = redeploy lever (see README).")
    ws.cell(row=row, column=1, value=legend).font = FONT_BODY

    _widths(ws, {"A": 30})
    for n in range(1, N_SPRINTS + 1):
        ws.column_dimensions[_col_letter(1 + n)].width = 6


def build_model_tracker(wb):
    ws = wb.create_sheet("Per-Model Tracker")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Per-Model Onboarding Tracker (10 models)"
    ws["A1"].font = FONT_TITLE

    cols = ["Model", "Group / Wave", "Sprint(s)", "SAS Owner", "Cadence"] + MODEL_STAGES
    _header(ws, 3, cols)

    row = 4
    for m in MODELS:
        name, group, sprints, owner, cadence = m
        vals = [name, group, sprints, owner, cadence] + ["Planned"] * len(MODEL_STAGES)
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = FONT_BODY_BOLD if ci == 1 else FONT_BODY
            c.alignment = ALIGN_LEFT if ci <= 5 else ALIGN_CENTER
            c.border = BORDER_THIN
            if "Group 1" in group and ci <= 5:
                c.fill = FILL_MILESTONE
        row += 1

    ws.cell(row=row + 1, column=1,
            value="Stages map to backlog tasks. Group 1 (models 1-2) is the proof; Group 2 (3-10) uses the "
                  "template, 2 models/sprint split across SAS + SAS2. Update each stage cell to Done/WIP/Blocked.").font = FONT_BODY

    _widths(ws, {"A": 10, "B": 14, "C": 10, "D": 10, "E": 12})
    for i in range(len(MODEL_STAGES)):
        ws.column_dimensions[_col_letter(6 + i)].width = 11


def build_governance(wb):
    ws = wb.create_sheet("Governance & Cadence")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Governance & Cadence"
    ws["A1"].font = FONT_TITLE

    ws["A3"] = "Ceremonies"
    ws["A3"].font = FONT_BODY_BOLD
    _header(ws, 4, ["Ceremony", "When", "Who", "Outcome"])
    row = 5
    for c_ in GOV_CEREMONIES:
        for ci, v in enumerate(c_, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_THIN
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Decision & Sign-off Checkpoints").font = FONT_BODY_BOLD
    row += 1
    _header(ws, row, ["Checkpoint", "Sprint", "Target date", "Owner", "Gate criteria"])
    row += 1
    for cp, sprint, owner, gate in GOV_CHECKPOINTS:
        s, e = sprint_dates(sprint)
        vals = [cp, f"S{sprint}", e, owner, gate]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = FONT_BODY_BOLD if ("SIGN-OFF" in cp or "CAPACITY" in cp) else FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_THIN
            if isinstance(v, date):
                cell.number_format = "yyyy-mm-dd"
        if "SIGN-OFF" in cp or "CAPACITY" in cp:
            for ci in range(1, 6):
                ws.cell(row=row, column=ci).fill = FILL_MILESTONE
        row += 1

    _widths(ws, {"A": 42, "B": 8, "C": 13, "D": 14, "E": 50})


def build_raid(wb):
    ws = wb.create_sheet("RAID Log")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "RAID Log - Risks / Assumptions / Issues / Dependencies"
    ws["A1"].font = FONT_TITLE

    _header(ws, 3, ["Type", "ID", "Item", "Owner", "Sprint / Prob-Impact", "Status / Mitigation"])
    row = 4
    type_fill = {
        "Dependency": FILL_BLOCKED,
        "Risk": FILL_WARN,
        "Assumption": FILL_OK,
        "Issue": FILL_IDLE,
    }
    for rec in RAID:
        rtype = rec[0]
        for ci, v in enumerate(rec, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = BORDER_THIN
        ws.cell(row=row, column=1).fill = type_fill.get(rtype, FILL_IDLE)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Internal prerequisites (Sprint 0 / Sprint 1)").font = FONT_BODY_BOLD
    row += 1
    _header(ws, row, ["ID", "Item", "Owner", "When"])
    row += 1
    for p in PREREQS_INTERNAL:
        for ci, v in enumerate(p, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = BORDER_THIN
        row += 1

    _widths(ws, {"A": 12, "B": 9, "C": 54, "D": 12, "E": 22, "F": 56})


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Delivery Dashboard"
    ws["A1"].font = FONT_TITLE

    ws["A3"] = "Tasks by Status"
    ws["A3"].font = FONT_BODY_BOLD
    _header(ws, 4, ["Status", "Count"])
    statuses = ["Planned", "In Progress", "Done", "Blocked"]
    row = 5
    for s in statuses:
        ws.cell(row=row, column=1, value=s).font = FONT_BODY
        ws.cell(row=row, column=2, value=f'=COUNTIFS(Backlog!$M:$M,"{s}",Backlog!$A:$A,"Task")')
        for ci in (1, 2):
            ws.cell(row=row, column=ci).border = BORDER_THIN
        row += 1

    ws["D3"] = "Effort by Role (d)"
    ws["D3"].font = FONT_BODY_BOLD
    for i, label in enumerate(["Role", "Days"], start=4):
        c = ws.cell(row=4, column=i, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER_THIN
    row = 5
    for role in ROLES:
        ws.cell(row=row, column=4, value=role).font = FONT_BODY
        ws.cell(row=row, column=5, value=f'=SUMIFS(Backlog!$H:$H,Backlog!$F:$F,"{role}",Backlog!$A:$A,"Task")').font = FONT_BODY
        for ci in (4, 5):
            ws.cell(row=row, column=ci).border = BORDER_THIN
        row += 1

    ws["A13"] = "Tasks + Effort by Epic"
    ws["A13"].font = FONT_BODY_BOLD
    _header(ws, 14, ["Epic", "Title", "Tasks", "Effort (d)", "Done"])
    row = 15
    for eid, etitle, _, _, _ in EPICS:
        ws.cell(row=row, column=1, value=eid).font = FONT_BODY_BOLD
        ws.cell(row=row, column=2, value=etitle).font = FONT_BODY
        ws.cell(row=row, column=3, value=f'=COUNTIFS(Backlog!$D:$D,"{eid}",Backlog!$A:$A,"Task")')
        ws.cell(row=row, column=4, value=f'=SUMIFS(Backlog!$H:$H,Backlog!$D:$D,"{eid}",Backlog!$A:$A,"Task")')
        ws.cell(row=row, column=5, value=f'=COUNTIFS(Backlog!$D:$D,"{eid}",Backlog!$A:$A,"Task",Backlog!$M:$M,"Done")')
        for ci in range(1, 6):
            ws.cell(row=row, column=ci).border = BORDER_THIN
        row += 1

    _widths(ws, {"A": 9, "B": 44, "C": 8, "D": 10, "E": 8, "F": 5, "G": 10, "H": 8})


def main():
    warnings = validate_capacity()
    print("Per-role load (d/sprint):")
    header = "role   " + " ".join(f"S{n:<4}" for n in range(1, N_SPRINTS + 1))
    print(header)
    load: dict[tuple[str, int], float] = {}
    for t in TASKS:
        load[(t[3], t[4])] = load.get((t[3], t[4]), 0.0) + t[5]
    for role in ROLES:
        cells = []
        for n in range(1, N_SPRINTS + 1):
            cells.append(f"{load.get((role, n), 0.0):<5.1f}")
        print(f"{role:<6} " + " ".join(cells))
    if warnings:
        print("\nCAPACITY WARNINGS:")
        for w in warnings:
            print(" ", w)
        raise SystemExit(1)

    repo_root = Path(__file__).resolve().parent.parent
    out = repo_root / "docs" / "absa-exl-agile-plan.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    build_backlog(wb)
    build_sprint_plan(wb)
    build_role_load(wb)
    build_model_tracker(wb)
    build_governance(wb)
    build_raid(wb)
    build_dashboard(wb)
    wb.save(out)

    n_tasks = len(TASKS)
    print(f"\nWrote {out}")
    print(f"  {len(EPICS)} epics, {len(STORIES)} stories, {n_tasks} tasks, "
          f"{sum(t[5] for t in TASKS):.1f} effort-days across {N_SPRINTS} sprints")


if __name__ == "__main__":
    main()
